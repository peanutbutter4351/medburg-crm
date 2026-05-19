"""
core/importers/base_importer.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Reusable foundation for all Excel import operations in Medburg CRM.

Architecture
────────────
┌─────────────────────────────────────────────────────────┐
│  BaseExcelImporter  (abstract)                          │
│                                                         │
│  ┌──────────────────┐    ┌────────────────────────────┐ │
│  │  _load_workbook  │    │  validate_headers          │ │
│  │  _get_sheet      │    │  _validate_row  (abstract) │ │
│  │  _iter_rows      │    │  import_file               │ │
│  └──────────────────┘    └────────────────────────────┘ │
│                                    │                    │
│                           returns  ▼                    │
│                          ImportResult                   │
└─────────────────────────────────────────────────────────┘

Usage (concrete importer):
    class DoctorImporter(BaseExcelImporter):
        REQUIRED_HEADERS = ["Name", "Mode", "City"]

        def _validate_row(self, row_num, row):
            errors = []
            name = self._cell(row, "Name")
            if not name:
                errors.append(self._err(row_num, "Name", "Name is required"))
            return errors

        def _import_row(self, row_num, row, result):
            # write to DB here
            ...

No business logic lives in this module.
"""

from __future__ import annotations

import logging
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger("medburg.importers")


# ── Result object ─────────────────────────────────────────────────────────────

@dataclass
class RowError:
    """
    A single validation or import error tied to a specific row and column.

    Attributes
    ----------
    row:
        1-based row number in the source Excel sheet (matches what a user
        sees in Excel — row 1 is the header, row 2 is the first data row).
    column:
        Human-readable column/field name (e.g. "Doctor Name", "Mode").
        Use the header label, not the column index, so error messages are
        actionable without the user counting columns.
    message:
        Plain-English description of what went wrong.
    """

    row: int
    column: str
    message: str

    def __str__(self) -> str:
        return f"Row {self.row} [{self.column}]: {self.message}"


@dataclass
class ImportResult:
    """
    Accumulated outcome of a single import operation.

    Passed into _import_row() so each concrete importer can update it
    in-place as rows succeed or fail.  Returned to the caller after
    import_file() completes.

    Attributes
    ----------
    total_rows:
        Number of data rows seen (excluding the header row).
    imported:
        Number of rows successfully written to the database.
    skipped:
        Number of rows intentionally skipped (e.g. duplicate detection).
    errors:
        List of RowError objects — one entry per validation failure.
        A single row can produce multiple errors (field-level granularity).
    """

    total_rows: int = 0
    imported: int = 0
    skipped: int = 0
    errors: list[RowError] = field(default_factory=list)

    # ── Convenience helpers ───────────────────────────────────────────────────

    @property
    def error_count(self) -> int:
        return len(self.errors)

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)

    @property
    def success_rate(self) -> float:
        """Fraction of rows imported successfully (0.0–1.0)."""
        if self.total_rows == 0:
            return 0.0
        return self.imported / self.total_rows

    def add_error(self, row: int, column: str, message: str) -> None:
        """Append a RowError without requiring callers to import RowError."""
        self.errors.append(RowError(row=row, column=column, message=message))

    def summary(self) -> str:
        """Single-line human-readable summary for logging and UI display."""
        return (
            f"Import complete — "
            f"{self.imported} imported, "
            f"{self.skipped} skipped, "
            f"{self.error_count} errors "
            f"(of {self.total_rows} total rows)"
        )

    def __repr__(self) -> str:  # pragma: no cover
        return (
            f"ImportResult("
            f"total={self.total_rows}, "
            f"imported={self.imported}, "
            f"skipped={self.skipped}, "
            f"errors={self.error_count})"
        )


# ── Base importer ─────────────────────────────────────────────────────────────

class BaseExcelImporter(ABC):
    """
    Abstract base class for all Excel-based import operations.

    Subclasses must declare:
        REQUIRED_HEADERS: list[str]
            Column headers that MUST be present in the uploaded file.
            Checked before any row is processed.

    Subclasses must implement:
        _validate_row(row_num, row) -> list[RowError]
            Business-rule validation for a single data row.
            Return an empty list if the row is valid.

        _import_row(row_num, row, result) -> None
            Persist a validated row (write to DB).
            Increment result.imported on success, result.skipped on skip.

    Optional override:
        SHEET_INDEX: int = 0
            Which sheet to read (0-based). Override if the target data
            is not on the first sheet.

        MAX_ROWS: int | None = None
            Hard cap on rows processed. Protects against runaway imports
            from malformed files. None = no limit.
    """

    # ── Subclass contract ─────────────────────────────────────────────────────

    #: Column headers that MUST exist in the uploaded file.
    REQUIRED_HEADERS: list[str] = []

    #: Sheet index (0-based) to read. Override for multi-sheet workbooks.
    SHEET_INDEX: int = 0

    #: Optional row cap — raises ImportError if exceeded.
    MAX_ROWS: int | None = None

    # ── Internal state ────────────────────────────────────────────────────────

    def __init__(self) -> None:
        self._header_map: dict[str, int] = {}  # header name → column index

    # ── Public entry point ────────────────────────────────────────────────────

    def import_file(self, file_path_or_buffer: Any) -> ImportResult:
        """
        Load an Excel file and process every data row.

        Parameters
        ----------
        file_path_or_buffer:
            Accepts a filesystem path (str / Path) OR a file-like object
            (e.g. Django's InMemoryUploadedFile, BytesIO).
            openpyxl handles both transparently.

        Returns
        -------
        ImportResult
            Fully populated result object. Callers decide what to do with
            errors — this method never raises on row-level failures.

        Raises
        ------
        ImportError
            On structural failures: unreadable file, wrong format,
            missing required headers, or MAX_ROWS exceeded.
        """
        result = ImportResult()

        try:
            wb = self._load_workbook(file_path_or_buffer)
            ws = self._get_sheet(wb)
        except Exception as exc:
            raise ImportError(f"Cannot read Excel file: {exc}") from exc

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ImportError("The Excel file contains no data.")

        # Row 1 is the header row (index 0 in the list)
        header_row = rows[0]
        self._build_header_map(header_row)

        missing = self.validate_headers()
        if missing:
            raise ImportError(
                f"Missing required column(s): {', '.join(missing)}. "
                f"Found columns: {', '.join(str(h) for h in header_row if h)}."
            )

        data_rows = rows[1:]
        result.total_rows = len(data_rows)

        if self.MAX_ROWS is not None and result.total_rows > self.MAX_ROWS:
            raise ImportError(
                f"File contains {result.total_rows} data rows, "
                f"which exceeds the maximum of {self.MAX_ROWS}. "
                f"Please split the file into smaller batches."
            )

        logger.info(
            "Starting import: %s — %d data rows",
            self.__class__.__name__,
            result.total_rows,
        )

        for idx, raw_row in enumerate(data_rows):
            # Excel row number: header=1, first data row=2
            row_num = idx + 2
            row = self._map_row(raw_row)

            # Skip fully empty rows silently (common at end of exports)
            if all(v is None or str(v).strip() == "" for v in raw_row):
                result.total_rows -= 1
                continue

            row_errors = self._validate_row(row_num, row)
            if row_errors:
                result.errors.extend(row_errors)
                continue  # Do NOT attempt DB write for invalid rows

            try:
                self._import_row(row_num, row, result)
            except Exception as exc:  # noqa: BLE001
                # Catch unexpected DB/logic errors so one bad row
                # doesn't abort the entire import.
                logger.exception(
                    "Unexpected error on row %d in %s: %s",
                    row_num,
                    self.__class__.__name__,
                    exc,
                )
                result.add_error(row_num, "—", f"Unexpected error: {exc}")

        logger.info(
            "Import complete: %s — %s",
            self.__class__.__name__,
            result.summary(),
        )
        return result

    # ── Header utilities ──────────────────────────────────────────────────────

    def validate_headers(self) -> list[str]:
        """
        Return a list of REQUIRED_HEADERS that are absent from the file.

        An empty return value means all required headers are present.
        Comparison is case-insensitive and strips leading/trailing whitespace,
        so "Doctor Name " and "doctor name" both match "Doctor Name".
        """
        normalised = {h.lower().strip() for h in self._header_map}
        return [
            h for h in self.REQUIRED_HEADERS
            if h.lower().strip() not in normalised
        ]

    def _build_header_map(self, header_row: tuple) -> None:
        """
        Build self._header_map from the raw header row tuple.

        Maps each non-empty header value (stripped, original case preserved)
        to its 0-based column index.
        """
        self._header_map = {
            str(cell).strip(): idx
            for idx, cell in enumerate(header_row)
            if cell is not None and str(cell).strip()
        }

    # ── Row utilities ─────────────────────────────────────────────────────────

    def _map_row(self, raw_row: tuple) -> dict[str, Any]:
        """
        Convert a raw openpyxl values-only row tuple into a header-keyed dict.

        Example
        -------
        header_map = {"Name": 0, "Mode": 1, "City": 2}
        raw_row    = ("Dr. Smith", "prepaid", "Mumbai")
        returns    → {"Name": "Dr. Smith", "Mode": "prepaid", "City": "Mumbai"}
        """
        return {
            header: (raw_row[idx] if idx < len(raw_row) else None)
            for header, idx in self._header_map.items()
        }

    def _cell(self, row: dict[str, Any], column: str) -> str:
        """
        Return the stripped string value for a column, or "" if missing/None.

        Use this inside _validate_row() for safe, None-safe cell access.

        Example
        -------
        name = self._cell(row, "Doctor Name")
        if not name:
            errors.append(self._err(row_num, "Doctor Name", "Required"))
        """
        value = row.get(column)
        if value is None:
            return ""
        return str(value).strip()

    def _cell_int(self, row: dict[str, Any], column: str) -> int | None:
        """
        Return the cell value as int, or None if blank or non-numeric.

        Useful for ID/quantity columns where None means "not provided".
        """
        raw = self._cell(row, column)
        if not raw:
            return None
        try:
            return int(float(raw))  # float() handles "42.0" from Excel
        except (ValueError, TypeError):
            return None

    def _cell_float(self, row: dict[str, Any], column: str) -> float | None:
        """
        Return the cell value as float, or None if blank or non-numeric.

        Useful for price/amount columns.
        """
        raw = self._cell(row, column)
        if not raw:
            return None
        try:
            return float(raw)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _err(row: int, column: str, message: str) -> RowError:
        """
        Convenience factory for RowError — keeps concrete importer code terse.

        Example
        -------
        errors.append(self._err(row_num, "Mode", "Must be 'prepaid' or 'postpaid'"))
        """
        return RowError(row=row, column=column, message=message)

    # ── File / sheet loading ──────────────────────────────────────────────────

    def _load_workbook(self, source: Any) -> openpyxl.Workbook:
        """
        Load the workbook with read_only=True and data_only=True.

        read_only:  streams the file without building the full in-memory
                    cell tree — critical for large imports (10k+ rows).
        data_only:  returns cell values, not formula strings.
                    Without this, price cells with SUM() formulas return
                    the formula text instead of the calculated number.
        """
        return openpyxl.load_workbook(
            source,
            read_only=True,
            data_only=True,
        )

    def _get_sheet(self, wb: openpyxl.Workbook) -> Worksheet:
        """
        Return the target worksheet by SHEET_INDEX.

        Raises ImportError with a clear message if the index is out of range.
        """
        sheet_names = wb.sheetnames
        if self.SHEET_INDEX >= len(sheet_names):
            raise ImportError(
                f"SHEET_INDEX={self.SHEET_INDEX} is out of range. "
                f"Workbook has {len(sheet_names)} sheet(s): {sheet_names}."
            )
        return wb[sheet_names[self.SHEET_INDEX]]

    # ── Abstract interface ────────────────────────────────────────────────────

    @abstractmethod
    def _validate_row(
        self,
        row_num: int,
        row: dict[str, Any],
    ) -> list[RowError]:
        """
        Validate a single data row against business rules.

        Parameters
        ----------
        row_num:
            1-based Excel row number (header=1, first data=2).
            Use in RowError so users can locate the problem in their file.
        row:
            Header-keyed dict — use self._cell(row, "Column Name") for
            safe string access, self._cell_int() / self._cell_float()
            for numeric columns.

        Returns
        -------
        list[RowError]
            Empty list → row is valid, proceed to _import_row().
            Non-empty  → row is skipped, errors added to ImportResult.
        """
        ...  # pragma: no cover

    @abstractmethod
    def _import_row(
        self,
        row_num: int,
        row: dict[str, Any],
        result: ImportResult,
    ) -> None:
        """
        Persist a single validated row to the database.

        Called ONLY when _validate_row() returns an empty error list.
        Must update result.imported or result.skipped before returning.

        Parameters
        ----------
        row_num:
            1-based Excel row number (for error reporting if a DB write fails).
        row:
            Header-keyed dict (same as _validate_row receives).
        result:
            Mutable ImportResult — update in-place:
                result.imported += 1   on successful DB write
                result.skipped  += 1   on intentional duplicate skip
        """
        ...  # pragma: no cover
