"""
Postpaid Doctors Report service layer.

Handles database-level aggregation using correlated subqueries, summary statistics,
and standardized Excel export rendering.
"""

import datetime
from decimal import Decimal
from io import BytesIO

from django.db.models import (
    DecimalField, OuterRef, Subquery, Sum, Value, Q,
)
from django.db.models.functions import Coalesce

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from doctors.models import Doctor
from medicines.models import Medicine
from sales.models import PostpaidCampaign, PostpaidSaleEntry


def get_postpaid_doctor_report_filter_options():
    """
    Return distinct filter dropdown options for the Postpaid Doctors Report.
    """
    doctors = (
        Doctor.objects
        .filter(is_active=True, mode="postpaid")
        .order_by("name")
        .values("id", "name")
    )
    
    locations = (
        Doctor.objects
        .filter(is_active=True, mode="postpaid")
        .exclude(location="")
        .exclude(location__isnull=True)
        .order_by("location")
        .values_list("location", flat=True)
        .distinct()
    )
    
    medicines = (
        Medicine.objects
        .filter(is_active=True)
        .order_by("name")
        .values("id", "name", "brand")
    )

    return {
        "doctors": list(doctors),
        "locations": list(locations),
        "medicines": list(medicines),
        "statuses": PostpaidCampaign.STATUS_CHOICES,
    }


def get_postpaid_doctor_report_queryset(
    *,
    doctor_id=None,
    location=None,
    medicine_id=None,
    from_date=None,
    to_date=None,
    status=None,
):
    """
    Return a Doctor queryset annotated with:
    - total_sales_value (Sum of PostpaidSaleEntry.value_at_sale)
    - total_commission (Sum of PostpaidSaleEntry.commission_at_sale)

    Filters are pushed down into the correlated subqueries to ensure correct
    re-calculation when status, products, or date ranges are filtered.
    """
    # Base queryset: active postpaid doctors with campaigns
    qs = Doctor.objects.filter(is_active=True, mode="postpaid", postpaid_campaigns__isnull=False).distinct()

    # Apply doctor-level filters
    if doctor_id:
        qs = qs.filter(id=doctor_id)
    if location:
        qs = qs.filter(location__iexact=location)

    # Filter doctors based on matching campaign/sale entry criteria
    if status:
        qs = qs.filter(postpaid_campaigns__status=status)
    if medicine_id:
        qs = qs.filter(postpaid_campaigns__sales_entries__medicine_id=medicine_id)
    if from_date:
        qs = qs.filter(postpaid_campaigns__sales_entries__entry_date__gte=from_date)
    if to_date:
        qs = qs.filter(postpaid_campaigns__sales_entries__entry_date__lte=to_date)

    qs = qs.distinct()

    # ── Correlated Subquery Filters ──
    sales_filters = Q(campaign__doctor_id=OuterRef("pk"))
    if status:
        sales_filters &= Q(campaign__status=status)
    if medicine_id:
        sales_filters &= Q(medicine_id=medicine_id)
    if from_date:
        sales_filters &= Q(entry_date__gte=from_date)
    if to_date:
        sales_filters &= Q(entry_date__lte=to_date)

    # ── Subquery 1: Sales Value ──
    sales_subquery = (
        PostpaidSaleEntry.objects
        .filter(sales_filters)
        .values("campaign__doctor_id")
        .annotate(total=Sum("value_at_sale"))
        .values("total")[:1]
    )

    # ── Subquery 2: Commission Earned ──
    commission_subquery = (
        PostpaidSaleEntry.objects
        .filter(sales_filters)
        .values("campaign__doctor_id")
        .annotate(total=Sum("commission_at_sale"))
        .values("total")[:1]
    )

    qs = qs.annotate(
        total_sales_value=Coalesce(
            Subquery(sales_subquery, output_field=DecimalField()),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        ),
        total_commission=Coalesce(
            Subquery(commission_subquery, output_field=DecimalField()),
            Value(Decimal("0.00")),
            output_field=DecimalField()
        )
    )

    return qs.order_by("name")


def get_postpaid_doctor_report_summary(queryset):
    """
    Aggregate totals across the filtered postpaid doctor queryset.
    
    Performs summation in Python via values_list to prevent SQLite nested aggregation bugs
    caused by outer-refs across joins.
    """
    total_doctors = queryset.count()
    total_sales = Decimal("0.00")
    total_commission = Decimal("0.00")
    
    for sales_val, comm_val in queryset.values_list("total_sales_value", "total_commission"):
        total_sales += sales_val or Decimal("0.00")
        total_commission += comm_val or Decimal("0.00")
        
    if total_sales > 0:
        average_commission_pct = (total_commission / total_sales) * 100
    else:
        average_commission_pct = Decimal("0.00")
        
    return {
        "total_doctors": total_doctors,
        "total_sales": total_sales,
        "total_commission": total_commission,
        "average_commission_pct": average_commission_pct,
    }


def export_postpaid_doctor_report_to_excel(queryset, summary, filters_desc=None):
    """
    Generate a styled Excel spreadsheet matching the filtered report structure.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Postpaid Doctors Report"

    # Style definitions
    font_company = Font(name="Calibri", bold=True, size=16, color="1E293B")
    font_title = Font(name="Calibri", bold=True, size=14, color="475569")
    font_meta = Font(name="Calibri", size=10, italic=True, color="64748B")
    font_filter_label = Font(name="Calibri", bold=True, size=10, color="334155")
    font_filter_val = Font(name="Calibri", size=10, color="334155")
    
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    summary_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    summary_font = Font(name="Calibri", bold=True, size=11, color="0F172A")
    currency_fmt = "₹#,##0.00"

    # Write Standard Metadata Header
    ws.cell(row=1, column=1, value="Medburg Medical Products").font = font_company
    ws.cell(row=2, column=1, value="Postpaid Doctors Report").font = font_title
    
    now_str = datetime.datetime.now().strftime("%d-%b-%Y %I:%M %p")
    ws.cell(row=3, column=1, value=f"Generated: {now_str}").font = font_meta
    
    ws.cell(row=5, column=1, value="Applied Filters:").font = Font(name="Calibri", bold=True, size=11, color="1E293B")
    
    current_row = 6
    if filters_desc:
        for label, val in filters_desc.items():
            cell_lbl = ws.cell(row=current_row, column=1, value=f"{label}:")
            cell_lbl.font = font_filter_label
            cell_val = ws.cell(row=current_row, column=2, value=str(val))
            cell_val.font = font_filter_val
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="None").font = font_filter_val
        current_row += 1
        
    current_row += 1  # Spacing row
    
    # Table headers
    columns = [
        ("Sl No", 10),
        ("Doctor Name", 30),
        ("Total Sales Value", 22),
        ("Total Commission", 22),
    ]
    
    header_row = current_row
    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 26
    
    current_row += 1
    
    # Table body
    c_align = Alignment(horizontal="center", vertical="center")
    r_align = Alignment(horizontal="right", vertical="center")
    l_align = Alignment(horizontal="left", vertical="center")
    
    for idx, doctor in enumerate(queryset, start=1):
        ws.cell(row=current_row, column=1, value=idx).alignment = c_align
        ws.cell(row=current_row, column=2, value=doctor.name).alignment = l_align
        
        sales_cell = ws.cell(row=current_row, column=3, value=float(doctor.total_sales_value))
        sales_cell.number_format = currency_fmt
        sales_cell.alignment = r_align
        
        comm_cell = ws.cell(row=current_row, column=4, value=float(doctor.total_commission))
        comm_cell.number_format = currency_fmt
        comm_cell.alignment = r_align
        
        is_alt = (idx % 2 == 0)
        for col_idx in range(1, 5):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_row_fill
        current_row += 1

    # Summary Row
    summary_row = current_row
    ws.cell(row=summary_row, column=1, value="Total").alignment = c_align
    ws.cell(row=summary_row, column=2, value=f"{summary['total_doctors']} Doctors").alignment = l_align
    
    tot_sales_cell = ws.cell(row=summary_row, column=3, value=float(summary["total_sales"]))
    tot_sales_cell.number_format = currency_fmt
    tot_sales_cell.alignment = r_align
    
    tot_comm_cell = ws.cell(row=summary_row, column=4, value=float(summary["total_commission"]))
    tot_comm_cell.number_format = currency_fmt
    tot_comm_cell.alignment = r_align
    
    for col_idx in range(1, 5):
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.font = summary_font
        cell.fill = summary_fill
        cell.border = thin_border
        
    current_row += 1
    
    # Average Commission % Row
    avg_row = current_row
    ws.merge_cells(start_row=avg_row, start_column=1, end_row=avg_row, end_column=3)
    ws.cell(row=avg_row, column=1, value="Average Commission %").alignment = Alignment(horizontal="right", vertical="center")
    
    avg_cell = ws.cell(row=avg_row, column=4, value=float(summary["average_commission_pct"])/100.0)
    avg_cell.number_format = "0.00%"
    avg_cell.alignment = r_align
    
    for col_idx in range(1, 5):
        cell = ws.cell(row=avg_row, column=col_idx)
        cell.font = summary_font
        cell.fill = summary_fill
        cell.border = thin_border
        
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
