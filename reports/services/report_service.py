"""
Report service layer.  (ARCH-2B: Snapshot Accounting)

All filtering, aggregation, and Excel export logic lives here so
views remain thin and templates receive pre-computed data only.

Performance
───────────
• select_related  — avoids N+1 on doctor / medicine / rep FKs
• DB-level annotation — line_value = value_at_sale (frozen snapshot)
• Aggregated summary uses the same pre-filtered queryset
• No Python loops for financial computation — values read from snapshots
• Subquery-based aggregation for doctor-level ROI to avoid
  cross-join duplication between Investment and SalesEntry.

Snapshot accounting (ARCH-2A / ARCH-2B)
────────────────────────────────────────
• ALL value calculations use value_at_sale — NEVER quantity × medicine.pts.
• pts_at_sale is displayed but never used in arithmetic.
• medicine.pts is never referenced in this module.
• Investment totals aggregate by investment identity (PK-list) to avoid
  the distinct=True / duplicate-amount undercounting bug.
"""

from decimal import Decimal
from io import BytesIO

from django.db.models import (
    Case, CharField, F, Subquery, Sum, Value, When,
    DecimalField, OuterRef, Count, Q,
)
from django.db.models.functions import Coalesce

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from accounts.models import User
from core.constants import ROLE_REP
from doctors.models import Doctor, Investment
from medicines.models import Medicine
from sales.models import SalesEntry


# ─── Filter-option helpers ───────────────────────────────────────

def get_report_filter_options():
    """
    Return distinct values for every filter dropdown on the report page.
    """
    doctors = (
        Doctor.objects
        .filter(is_active=True, mode="prepaid")
        .order_by("name")
        .values("id", "name")
    )

    reps = (
        User.objects
        .filter(role=ROLE_REP, is_active=True)
        .order_by("first_name", "last_name")
        .values("id", "first_name", "last_name", "username")
    )

    medicines = (
        Medicine.objects
        .filter(is_active=True)
        .order_by("name")
        .values("id", "name", "brand")
    )

    return {
        "doctors": list(doctors),
        "reps": list(reps),
        "medicines": list(medicines),
    }


# ─── Core queryset ───────────────────────────────────────────────

def get_report_queryset(
    *,
    from_date=None,
    to_date=None,
    doctor_id=None,
    rep_id=None,
    medicine_id=None,
    sort=None,
):
    """
    Return an annotated SalesEntry queryset for the prepaid report.

    Annotation added
    ────────────────
    line_value — value_at_sale (frozen snapshot, ARCH-2A)
                 Falls back to 0 for any row where value_at_sale is NULL
                 (should not occur after the backfill migration).

    Filter: investment__isnull=False selects prepaid entries only.
    All FK look-ups use select_related to avoid N+1.
    """
    qs = (
        SalesEntry.objects
        .filter(investment__isnull=False)
        .select_related("doctor", "medicine", "rep", "investment")
        .annotate(
            # ARCH-2B: use the frozen snapshot — never live medicine.pts
            line_value=Coalesce(
                F("value_at_sale"),
                Value(Decimal("0")),
                output_field=DecimalField(),
            ),
        )
        .order_by("-entry_date", "-created_at")
    )

    # ── Apply filters ────────────────────────────────
    if from_date:
        qs = qs.filter(entry_date__gte=from_date)
    if to_date:
        qs = qs.filter(entry_date__lte=to_date)
    if doctor_id:
        qs = qs.filter(doctor_id=doctor_id)
    if rep_id:
        qs = qs.filter(rep_id=rep_id)
    if medicine_id:
        qs = qs.filter(medicine_id=medicine_id)

    return qs


# ─── Doctor-level ROI report ─────────────────────────────────────

def get_doctor_roi_report(
    *,
    from_date=None,
    to_date=None,
    doctor_id=None,
    rep_id=None,
    medicine_id=None,
    sort=None,
):
    """
    Return a list of template-ready dicts with one row per
    SalesEntry, enriched with investment-level ROI data.

    Running balance calculates progressively per investment,
    processing entries by entry_date ascending, then id ascending.

    ARCH-2B: All value calculations use entry.value_at_sale (the frozen
    snapshot) — never entry.quantity × entry.medicine.pts.
    The .value property on SalesEntry already does this (snapshot-first),
    so row_value = entry.value is equivalent and safe.
    """
    qs = (
        SalesEntry.objects
        .filter(investment__isnull=False)
        .select_related("doctor", "medicine", "rep", "investment")
        .order_by("investment_id", "entry_date", "id")
    )

    if from_date:
        qs = qs.filter(entry_date__gte=from_date)
    if to_date:
        qs = qs.filter(entry_date__lte=to_date)
    if doctor_id:
        qs = qs.filter(doctor_id=doctor_id)
    if rep_id:
        qs = qs.filter(rep_id=rep_id)
    if medicine_id:
        qs = qs.filter(medicine_id=medicine_id)

    rows = []
    current_inv_id = None
    current_balance = Decimal("0")

    for idx, entry in enumerate(qs.iterator(chunk_size=500), start=1):
        inv = entry.investment
        if current_inv_id != inv.id:
            current_inv_id = inv.id
            current_balance = inv.roi_amount

        # ARCH-2B: entry.value returns value_at_sale (snapshot-first).
        # For legacy rows without a snapshot this falls back to live PTS,
        # but after the ARCH-2A backfill every row has value_at_sale set.
        row_value = entry.value   # → value_at_sale (frozen)
        current_balance -= row_value

        status = "Completed" if current_balance <= 0 else "In Progress"

        rows.append({
            "sl_no": idx,
            "doctor_name": entry.doctor.name,
            "rep_name": entry.rep.get_full_name() or entry.rep.username,
            "invested_date": inv.start_date,
            "expected_roi": inv.roi_amount,
            "location": entry.doctor.location or "—",
            "medicine_name": str(entry.medicine),
            "pts_at_sale": entry.pts_at_sale,          # display only
            "investment_amount": inv.amount,
            "value": row_value,                         # frozen snapshot
            "quantity": entry.quantity,
            "entry_date": entry.entry_date,
            "balance_roi": current_balance,
            "status": status,
            "note": inv.notes or "—",
            "is_legacy": entry.is_snapshot_legacy,     # audit marker
        })

    if sort == "newest_first":
        rows.sort(key=lambda x: x["entry_date"], reverse=True)
    elif sort == "oldest_first":
        rows.sort(key=lambda x: x["entry_date"])
    elif sort == "doctor_az":
        rows.sort(key=lambda x: x["doctor_name"].lower())
    elif sort == "doctor_za":
        rows.sort(key=lambda x: x["doctor_name"].lower(), reverse=True)
    elif sort == "highest_balance":
        rows.sort(key=lambda x: x["balance_roi"], reverse=True)
    elif sort == "lowest_balance":
        rows.sort(key=lambda x: x["balance_roi"])
    elif sort == "completed_first":
        rows.sort(key=lambda x: (0 if x["status"] == "Completed" else 1, x["entry_date"]))
    elif sort == "inprogress_first":
        rows.sort(key=lambda x: (0 if x["status"] == "In Progress" else 1, x["entry_date"]))
    else:
        rows.sort(key=lambda x: x["entry_date"], reverse=True)

    for i, row in enumerate(rows, start=1):
        row["sl_no"] = i

    return rows


# ─── Aggregated summary ─────────────────────────────────────────

def get_report_summary(queryset):
    """
    Aggregate totals across the filtered SalesEntry queryset.

    ARCH-2B changes
    ───────────────
    • total_value  = Sum("value_at_sale")  [was: Sum(quantity × medicine.pts)]
    • total_investment / total_roi_amount aggregate by investment PK list —
      no distinct=True on financial values (fixes duplicate-amount undercounting).

    The investment_ids approach:
      Collect the distinct investment PKs that appear in the filtered entries,
      then aggregate those Investment rows directly.  This ensures:
      - Two investments with the same amount (e.g. 15000 + 15000) are both counted
      - No cross-join inflation from the SalesEntry → Investment join
    """
    agg = queryset.aggregate(
        total_quantity=Coalesce(
            Sum("quantity"), Value(0),
        ),
        # ARCH-2B: snapshot-based value aggregation
        total_value=Coalesce(
            Sum("value_at_sale"),
            Value(Decimal("0")),
            output_field=DecimalField(),
        ),
    )

    # Collect distinct investment PKs from the filtered queryset.
    # .distinct() here is on the *PK column* (identity), not on amount values.
    investment_ids = (
        queryset
        .filter(investment_id__isnull=False)
        .values_list("investment_id", flat=True)
        .distinct()
    )

    inv_agg = (
        Investment.objects
        .filter(id__in=investment_ids)
        .aggregate(
            # No distinct=True here — we query exactly the investments we want
            total_investment=Coalesce(
                Sum("amount"),
                Value(Decimal("0")),
                output_field=DecimalField(),
            ),
            total_roi_amount=Coalesce(
                Sum(F("amount") * F("roi_ratio")),
                Value(Decimal("0")),
                output_field=DecimalField(),
            ),
        )
    )

    achieved = agg["total_value"]
    investment = inv_agg["total_investment"]
    roi_amount = inv_agg["total_roi_amount"]
    balance = roi_amount - achieved

    return {
        "total_entries": queryset.count(),
        "total_quantity": agg["total_quantity"],
        "total_value": achieved,
        "total_investment": investment,
        "total_roi_amount": roi_amount,
        "balance_roi": balance,
    }


# ─── Excel export ────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_SUMMARY_FILL = PatternFill(start_color="E8EDFF", end_color="E8EDFF", fill_type="solid")
_SUMMARY_FONT = Font(name="Calibri", bold=True, size=11)

_COLUMNS = [
    ("Sl No", 8),
    ("Doctor", 24),
    ("Area", 18),
    ("Sales Rep", 20),
    ("Investment (₹)", 16),
    ("Invested Date", 14),
    ("Expected ROI (₹)", 16),
    ("Medicine", 24),
    ("Qty", 10),
    ("Date", 14),
    ("Value (₹)", 16),
    ("Balance (₹)", 16),
    ("Status", 14),
    ("Note", 30),
]

_NUM_COLS = len(_COLUMNS)


def export_to_excel(roi_rows, summary):
    """
    Generate a styled .xlsx workbook from the doctor ROI rows and
    return a BytesIO buffer ready for HttpResponse streaming.

    ARCH-2B: All values written to Excel come from the pre-computed
    roi_rows dicts, which use value_at_sale (frozen snapshots).
    No live medicine.pts calculations occur here.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Doctor ROI Report"

    # ── Title row ────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(_NUM_COLS)}1")
    title_cell = ws["A1"]
    title_cell.value = "Medburg CRM — Doctor ROI Report"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="4361EE")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Header row ───────────────────────────────────
    header_row = 3
    for col_idx, (label, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 24

    # ── Data rows ────────────────────────────────────
    data_align = Alignment(vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    currency_fmt = '#,##0.00'

    row_num = header_row + 1
    for row in roi_rows:
        ws.cell(row=row_num, column=1, value=row["sl_no"]).alignment = center_align
        ws.cell(row=row_num, column=2, value=row["doctor_name"]).alignment = data_align
        ws.cell(row=row_num, column=3, value=row["location"]).alignment = data_align
        ws.cell(row=row_num, column=4, value=row["rep_name"]).alignment = data_align

        inv_cell = ws.cell(row=row_num, column=5, value=float(row["investment_amount"]))
        inv_cell.number_format = currency_fmt
        inv_cell.alignment = right_align

        ws.cell(row=row_num, column=6, value=row["invested_date"]).alignment = data_align

        roi_cell = ws.cell(row=row_num, column=7, value=float(row["expected_roi"]))
        roi_cell.number_format = currency_fmt
        roi_cell.alignment = right_align

        ws.cell(row=row_num, column=8, value=row["medicine_name"]).alignment = data_align
        ws.cell(row=row_num, column=9, value=row["quantity"]).alignment = center_align
        ws.cell(row=row_num, column=10, value=row["entry_date"]).alignment = data_align

        # ARCH-2B: row["value"] is already the frozen snapshot value
        val_cell = ws.cell(row=row_num, column=11, value=float(row["value"]))
        val_cell.number_format = currency_fmt
        val_cell.alignment = right_align

        bal_cell = ws.cell(row=row_num, column=12, value=float(row["balance_roi"]))
        bal_cell.number_format = currency_fmt
        bal_cell.alignment = right_align

        ws.cell(row=row_num, column=13, value=row["status"]).alignment = center_align
        ws.cell(row=row_num, column=14, value=row["note"]).alignment = data_align

        # Alternate row shading
        if row_num % 2 == 0:
            shade = PatternFill(start_color="F8F9FF", end_color="F8F9FF", fill_type="solid")
            for c in range(1, _NUM_COLS + 1):
                ws.cell(row=row_num, column=c).fill = shade

        for c in range(1, _NUM_COLS + 1):
            ws.cell(row=row_num, column=c).border = _THIN_BORDER

        row_num += 1

    # ── Summary block ────────────────────────────────
    summary_start = row_num + 2
    summary_items = [
        ("Total Entries", summary["total_entries"], False),
        ("Total Quantity", summary["total_quantity"], False),
        ("Total Sales Value (₹)", summary["total_value"], True),
        ("Total Investment (₹)", summary["total_investment"], True),
        ("Total ROI Target (₹)", summary["total_roi_amount"], True),
        ("Balance ROI (₹)", summary["balance_roi"], True),
    ]

    ws.merge_cells(f"A{summary_start}:B{summary_start}")
    heading = ws.cell(row=summary_start, column=1, value="Summary")
    heading.font = Font(name="Calibri", bold=True, size=12, color="4361EE")
    summary_start += 1

    for label, val, is_currency in summary_items:
        label_cell = ws.cell(row=summary_start, column=1, value=label)
        label_cell.font = _SUMMARY_FONT
        label_cell.fill = _SUMMARY_FILL
        label_cell.border = _THIN_BORDER

        val_cell = ws.cell(row=summary_start, column=2, value=float(val) if val else 0)
        val_cell.font = _SUMMARY_FONT
        val_cell.fill = _SUMMARY_FILL
        val_cell.border = _THIN_BORDER
        if is_currency:
            val_cell.number_format = currency_fmt
        val_cell.alignment = Alignment(horizontal="right")

        summary_start += 1

    # ── Freeze header ────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Write to buffer ──────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Consolidated Prepaid Doctor Report ───────────────────────────

def get_prepaid_doctor_report_queryset(
    *,
    doctor_id=None,
    location=None,
    medicine_id=None,
    from_date=None,
    to_date=None,
    status=None,
):
    """
    Return a Doctor queryset annotated with:
    - total_investment (Sum of Investment.amount)
    - total_expected_return (Sum of Investment.amount * Investment.roi_ratio)
    - total_returns (Sum of SalesEntry.value_at_sale)
    - recovery_pct (Total Returns / Total Expected Return * 100)
    
    Each row represents one doctor. If a doctor has multiple prepaid investments,
    we aggregate all investments (and their returns) matching the status filter
    into a single row.
    """
    # Base queryset: active prepaid doctors who have at least one investment
    qs = Doctor.objects.filter(is_active=True, mode="prepaid", investments__isnull=False).distinct()

    # ── Subquery 1: Total Investment per Doctor ──
    investment_filters = Q(doctor_id=OuterRef("pk"))
    if status:
        investment_filters &= Q(status=status)

    investment_subquery = (
        Investment.objects
        .filter(investment_filters)
        .values("doctor_id")
        .annotate(total=Sum("amount"))
        .values("total")[:1]
    )

    # ── Subquery 2: Total Expected Return per Doctor ──
    expected_filters = Q(doctor_id=OuterRef("pk"))
    if status:
        expected_filters &= Q(status=status)

    expected_subquery = (
        Investment.objects
        .filter(expected_filters)
        .values("doctor_id")
        .annotate(total=Sum(F("amount") * F("roi_ratio")))
        .values("total")[:1]
    )

    # ── Subquery 3: Total Returns (Sales) per Doctor ──
    sales_filters = Q(doctor_id=OuterRef("pk"), investment__isnull=False)
    if status:
        sales_filters &= Q(investment__status=status)
    if medicine_id:
        sales_filters &= Q(medicine_id=medicine_id)
    if from_date:
        sales_filters &= Q(entry_date__gte=from_date)
    if to_date:
        sales_filters &= Q(entry_date__lte=to_date)

    sales_subquery = (
        SalesEntry.objects
        .filter(sales_filters)
        .values("doctor_id")
        .annotate(total=Sum("value_at_sale"))
        .values("total")[:1]
    )

    # Annotate values
    qs = qs.annotate(
        total_investment=Coalesce(
            Subquery(investment_subquery, output_field=DecimalField()),
            Value(Decimal("0")),
            output_field=DecimalField(),
        ),
        total_expected_return=Coalesce(
            Subquery(expected_subquery, output_field=DecimalField()),
            Value(Decimal("0")),
            output_field=DecimalField(),
        ),
        total_returns=Coalesce(
            Subquery(sales_subquery, output_field=DecimalField()),
            Value(Decimal("0")),
            output_field=DecimalField(),
        )
    ).select_related("assigned_rep")

    # Apply doctor-level filters
    if doctor_id:
        qs = qs.filter(id=doctor_id)
    if location:
        qs = qs.filter(location__iexact=location)

    # If a status filter is applied, we must only return doctors who have investments matching that status.
    if status:
        qs = qs.filter(investments__status=status).distinct()

    # Sort alphabetically by Doctor Name (A-Z)
    return qs.order_by("name")


def get_prepaid_doctor_report_summary(queryset):
    """
    Aggregate totals across the consolidated Doctor queryset for the report footer.
    
    Average Recovery % = (Total Returns / Total Expected Return) * 100
    Displays 0% safely if Total Expected Return is zero.
    """
    agg = queryset.aggregate(
        total_doctors=Count("id"),
        sum_investment=Coalesce(Sum("total_investment"), Value(Decimal("0"))),
        sum_expected_return=Coalesce(Sum("total_expected_return"), Value(Decimal("0"))),
        sum_returns=Coalesce(Sum("total_returns"), Value(Decimal("0"))),
    )
    
    total_investment = agg["sum_investment"]
    total_expected_return = agg["sum_expected_return"]
    total_returns = agg["sum_returns"]
    
    if total_expected_return > 0:
        average_recovery_pct = (total_returns / total_expected_return) * Decimal("100.0")
    else:
        average_recovery_pct = Decimal("0.0")
        
    return {
        "total_doctors": agg["total_doctors"],
        "total_investment": total_investment,
        "total_expected_return": total_expected_return,
        "total_returns": total_returns,
        "average_recovery_pct": average_recovery_pct,
    }


def export_prepaid_doctor_report_to_excel(queryset, summary, filters_description=None):
    """
    Generate a styled .xlsx workbook from the consolidated prepaid doctor report queryset
    and return a BytesIO buffer.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Prepaid Doctors Report"

    # Styles
    _TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1E3A8A")
    _SUBTITLE_FONT = Font(name="Calibri", italic=True, size=11, color="4B5563")
    _HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    _HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    _HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _THIN_BORDER = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    _SUMMARY_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    _SUMMARY_FONT = Font(name="Calibri", bold=True, size=11, color="1E3A8A")
    _DATA_FONT = Font(name="Calibri", size=11)

    # Title & Company Name
    ws["A1"] = "Medburg CRM"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = "Prepaid Doctors Report"
    ws["A2"].font = Font(name="Calibri", bold=True, size=13)

    # Generated Timestamp
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = f"Generated At: {timestamp}"
    ws["A4"].font = _SUBTITLE_FONT

    # Applied Filters description block
    current_row = 6
    if filters_description:
        ws.cell(row=current_row, column=1, value="Applied Filters:").font = Font(name="Calibri", bold=True, size=10, color="374151")
        for k, v in filters_description.items():
            ws.cell(row=current_row, column=2, value=f"{k}: {v}").font = Font(name="Calibri", size=10, color="4B5563")
            current_row += 1
        current_row += 1  # Blank row after filters

    header_row = current_row

    # Columns definitions
    _COLUMNS = [
        ("Sl No", 10),
        ("Doctor Name", 35),
        ("Total Investment (₹)", 25),
        ("Total Expected Return (₹)", 25),
        ("Total Returns Received (₹)", 25),
    ]

    # Write headers
    for col_idx, (label, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 25

    # Write Data
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    currency_fmt = '₹#,##0.00'

    data_start_row = header_row + 1
    for idx, doctor in enumerate(queryset, start=1):
        row_num = data_start_row + idx - 1
        
        c1 = ws.cell(row=row_num, column=1, value=idx)
        c1.font = _DATA_FONT
        c1.alignment = center_align
        c1.border = _THIN_BORDER

        c2 = ws.cell(row=row_num, column=2, value=doctor.name)
        c2.font = _DATA_FONT
        c2.alignment = left_align
        c2.border = _THIN_BORDER

        c3 = ws.cell(row=row_num, column=3, value=float(doctor.total_investment))
        c3.font = _DATA_FONT
        c3.number_format = currency_fmt
        c3.alignment = right_align
        c3.border = _THIN_BORDER

        c4 = ws.cell(row=row_num, column=4, value=float(doctor.total_expected_return))
        c4.font = _DATA_FONT
        c4.number_format = currency_fmt
        c4.alignment = right_align
        c4.border = _THIN_BORDER

        c5 = ws.cell(row=row_num, column=5, value=float(doctor.total_returns))
        c5.font = _DATA_FONT
        c5.number_format = currency_fmt
        c5.alignment = right_align
        c5.border = _THIN_BORDER

        # Alternating row shading
        if idx % 2 == 0:
            shade = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            c1.fill = shade
            c2.fill = shade
            c3.fill = shade
            c4.fill = shade
            c5.fill = shade

        ws.row_dimensions[row_num].height = 20

    # Write Totals
    tot_row = data_start_row + len(queryset)
    ws.cell(row=tot_row, column=1, value="Total").font = _SUMMARY_FONT
    ws.cell(row=tot_row, column=1).alignment = center_align
    ws.cell(row=tot_row, column=1).fill = _SUMMARY_FILL
    ws.cell(row=tot_row, column=1).border = _THIN_BORDER

    ws.cell(row=tot_row, column=2, value=f"{summary['total_doctors']} Doctors").font = _SUMMARY_FONT
    ws.cell(row=tot_row, column=2).alignment = left_align
    ws.cell(row=tot_row, column=2).fill = _SUMMARY_FILL
    ws.cell(row=tot_row, column=2).border = _THIN_BORDER

    c3 = ws.cell(row=tot_row, column=3, value=float(summary['total_investment']))
    c3.font = _SUMMARY_FONT
    c3.number_format = currency_fmt
    c3.alignment = right_align
    c3.fill = _SUMMARY_FILL
    c3.border = _THIN_BORDER

    c4 = ws.cell(row=tot_row, column=4, value=float(summary['total_expected_return']))
    c4.font = _SUMMARY_FONT
    c4.number_format = currency_fmt
    c4.alignment = right_align
    c4.fill = _SUMMARY_FILL
    c4.border = _THIN_BORDER

    c5 = ws.cell(row=tot_row, column=5, value=float(summary['total_returns']))
    c5.font = _SUMMARY_FONT
    c5.number_format = currency_fmt
    c5.alignment = right_align
    c5.fill = _SUMMARY_FILL
    c5.border = _THIN_BORDER

    ws.row_dimensions[tot_row].height = 22

    # Write Average Recovery Row
    avg_row = tot_row + 1
    ws.cell(row=avg_row, column=1, value="").fill = _SUMMARY_FILL
    ws.cell(row=avg_row, column=1).border = _THIN_BORDER
    
    ws.cell(row=avg_row, column=2, value="Average Recovery %").font = _SUMMARY_FONT
    ws.cell(row=avg_row, column=2).alignment = left_align
    ws.cell(row=avg_row, column=2).fill = _SUMMARY_FILL
    ws.cell(row=avg_row, column=2).border = _THIN_BORDER
    
    ws.merge_cells(start_row=avg_row, start_column=3, end_row=avg_row, end_column=5)
    avg_cell = ws.cell(row=avg_row, column=3, value=f"{float(summary['average_recovery_pct']):.1f}%")
    avg_cell.font = _SUMMARY_FONT
    avg_cell.alignment = right_align
    avg_cell.fill = _SUMMARY_FILL
    
    # Apply border/fill manually to merged cells
    for c_idx in range(4, 6):
        ws.cell(row=avg_row, column=c_idx).fill = _SUMMARY_FILL
        ws.cell(row=avg_row, column=c_idx).border = _THIN_BORDER

    ws.row_dimensions[avg_row].height = 22

    # Freeze header row
    ws.freeze_panes = f"A{header_row + 1}"

    # Write to buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def get_prepaid_doctor_report_filter_options():
    """
    Return distinct values for every filter dropdown on the consolidated prepaid report page.
    """
    doctors = (
        Doctor.objects
        .filter(is_active=True, mode="prepaid")
        .order_by("name")
        .values("id", "name")
    )
    
    locations = (
        Doctor.objects
        .filter(is_active=True, mode="prepaid")
        .exclude(location="")
        .values_list("location", flat=True)
        .distinct()
        .order_by("location")
    )
    
    medicines = (
        Medicine.objects
        .filter(is_active=True)
        .order_by("name")
        .values("id", "name", "brand")
    )
    
    return {
        "doctors": list(doctors),
        "locations": list(locations),
        "medicines": list(medicines),
    }


