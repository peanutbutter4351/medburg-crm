"""
Postpaid campaign service layer — MR-9.0 Report Separation.

Provides:
  - get_campaign_queryset / get_campaign_summary / get_campaign_filter_options
    (used by the existing postpaid_report management view)

  - get_postpaid_sales_report / get_postpaid_sales_summary
    (Postpaid Sales Report: PostpaidSaleEntry-level rows)

  - get_settlement_ledger_report / get_settlement_summary
    (Settlement Ledger Report: PostpaidCampaign-level financial ledger)

  - get_postpaid_report_filter_options
    (shared filter options for both new reports)

  - export_postpaid_sales_to_excel / export_settlement_ledger_to_excel
    (Excel export for both new reports)
"""

import calendar
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.db.models import Q, Sum, Value, DecimalField
from django.db.models.functions import Coalesce

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sales.models import PostpaidCampaign, PostpaidSaleEntry
from doctors.models import Doctor

User = get_user_model()


# ─────────────────────────────────────────────────────────────────────────────
# Shared Excel style constants
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL  = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_SUMMARY_FILL = PatternFill(start_color="E8EDFF", end_color="E8EDFF", fill_type="solid")
_SUMMARY_FONT = Font(name="Calibri", bold=True, size=11)
_CURRENCY_FMT = "#,##0.00"
_ALT_ROW_FILL = PatternFill(start_color="F8FAFF", end_color="F8FAFF", fill_type="solid")

def _apply_header_row(ws, columns, row_num):
    """Write styled header cells and return the next row number."""
    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=label)
        cell.font   = _HEADER_FONT
        cell.fill   = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row_num].height = 24
    return row_num + 1


def _write_summary_block(ws, summary_items, start_row):
    """Write a two-column summary block below the data table."""
    ws.cell(row=start_row, column=1, value="Summary").font = Font(
        name="Calibri", bold=True, size=12, color="4361EE"
    )
    start_row += 1
    for label, val, is_currency in summary_items:
        label_cell = ws.cell(row=start_row, column=1, value=label)
        label_cell.font  = _SUMMARY_FONT
        label_cell.fill  = _SUMMARY_FILL
        label_cell.border = _THIN_BORDER

        val_cell = ws.cell(row=start_row, column=2, value=float(val) if val else 0)
        val_cell.font  = _SUMMARY_FONT
        val_cell.fill  = _SUMMARY_FILL
        val_cell.border = _THIN_BORDER
        val_cell.alignment = Alignment(horizontal="right")
        if is_currency:
            val_cell.number_format = _CURRENCY_FMT
        start_row += 1


# ─────────────────────────────────────────────────────────────────────────────
# A. Existing campaign management helpers (used by postpaid_report_view)
# ─────────────────────────────────────────────────────────────────────────────

def get_campaign_queryset(*, month=None, year=None, doctor_id=None, status=None, search=None):
    """
    Return a PostpaidCampaign queryset with filters applied.
    Used by the existing campaign management / monitoring report.
    """
    qs = (
        PostpaidCampaign.objects
        .select_related("doctor")
        .prefetch_related("payments", "sales_entries")
        .order_by("-year", "-month", "doctor__name")
    )

    if month:
        qs = qs.filter(month=month)
    if year:
        qs = qs.filter(year=year)
    if doctor_id:
        qs = qs.filter(doctor_id=doctor_id)
    if status:
        qs = qs.filter(status=status)
    if search:
        qs = qs.filter(
            Q(doctor__name__icontains=search) |
            Q(doctor__hospital__icontains=search)
        )
    return qs


def get_campaign_summary(queryset):
    """Aggregate totals across the filtered PostpaidCampaign queryset."""
    agg = queryset.aggregate(
        sales=Coalesce(Sum("total_sales_value"), Value(Decimal("0")), output_field=DecimalField()),
        commission=Coalesce(Sum("total_commission"), Value(Decimal("0")), output_field=DecimalField()),
        paid=Coalesce(Sum("paid_amount"), Value(Decimal("0")), output_field=DecimalField()),
    )
    sales       = agg["sales"]
    commission  = agg["commission"]
    paid        = agg["paid"]
    outstanding = commission - paid
    return {
        "total_campaigns":  queryset.count(),
        "total_sales":      sales,
        "total_commission": commission,
        "total_paid":       paid,
        "total_outstanding": outstanding,
    }


def get_campaign_filter_options():
    """
    Return distinct values for filter dropdowns on the campaign management report.
    """
    doctors = (
        Doctor.objects
        .filter(is_active=True, mode="postpaid")
        .order_by("name")
        .values("id", "name")
    )
    campaign_years = (
        PostpaidCampaign.objects
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )
    years = list(campaign_years)
    if not years:
        import datetime
        years = [datetime.date.today().year]

    return {
        "doctors": list(doctors),
        "years": years,
        "statuses": PostpaidCampaign.STATUS_CHOICES,
    }


# ─────────────────────────────────────────────────────────────────────────────
# B. Shared filter options for the two new MR-9.0 reports
# ─────────────────────────────────────────────────────────────────────────────

def get_postpaid_report_filter_options():
    """
    Return filter dropdown options shared by the Postpaid Sales Report
    and the Settlement Ledger Report.
    """
    doctors = (
        Doctor.objects
        .filter(is_active=True, mode="postpaid")
        .order_by("name")
        .values("id", "name")
    )
    reps = (
        User.objects
        .filter(is_active=True)
        .exclude(role="admin")
        .order_by("first_name", "last_name", "username")
        .values("id", "first_name", "last_name", "username")
    )
    campaign_years = (
        PostpaidCampaign.objects
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )
    years = list(campaign_years)
    if not years:
        import datetime
        years = [datetime.date.today().year]

    months = [(i, calendar.month_name[i]) for i in range(1, 13)]

    return {
        "doctors": list(doctors),
        "reps": list(reps),
        "years": years,
        "months": months,
        "statuses": PostpaidCampaign.STATUS_CHOICES,
    }


# ─────────────────────────────────────────────────────────────────────────────
# C. Postpaid Sales Report — PostpaidSaleEntry level
# ─────────────────────────────────────────────────────────────────────────────

def get_postpaid_sales_queryset(
    *,
    doctor_id=None,
    rep_id=None,
    month=None,
    year=None,
    medicine_id=None,
    status=None,
    from_date=None,
    to_date=None,
):
    """
    Return a filtered PostpaidSaleEntry queryset.
    All values read from frozen snapshots (pts_at_sale, value_at_sale,
    commission_at_sale) — never from live medicine.pts.
    """
    qs = (
        PostpaidSaleEntry.objects
        .select_related(
            "campaign",
            "campaign__doctor",
            "medicine",
            "rep",
        )
        .order_by("-campaign__year", "-campaign__month", "campaign__doctor__name", "-entry_date")
    )

    if doctor_id:
        qs = qs.filter(campaign__doctor_id=doctor_id)
    if rep_id:
        qs = qs.filter(rep_id=rep_id)
    if month:
        qs = qs.filter(campaign__month=month)
    if year:
        qs = qs.filter(campaign__year=year)
    if medicine_id:
        qs = qs.filter(medicine_id=medicine_id)
    if status:
        qs = qs.filter(campaign__status=status)
    if from_date:
        qs = qs.filter(entry_date__gte=from_date)
    if to_date:
        qs = qs.filter(entry_date__lte=to_date)

    return qs


def get_postpaid_sales_report(qs):
    """
    Convert a PostpaidSaleEntry queryset to a list of template-ready dicts.
    All monetary values come from frozen snapshot fields.
    """
    rows = []
    for idx, entry in enumerate(qs.iterator(chunk_size=500), start=1):
        camp = entry.campaign
        rep  = entry.rep
        rows.append({
            "sl_no":             idx,
            "period":            f"{camp.month:02d}/{camp.year}",
            "month":             camp.month,
            "year":              camp.year,
            "doctor_name":       camp.doctor.name,
            "doctor_location":   getattr(camp.doctor, "location", "") or "—",
            "rep_name":          (rep.get_full_name() or rep.username) if rep else "—",
            "medicine_name":     str(entry.medicine),
            "quantity":          entry.quantity,
            "pts_at_sale":       entry.pts_at_sale,
            "value_at_sale":     entry.value_at_sale,
            "commission_pct":    entry.commission_percentage_at_sale,
            "commission_earned": entry.commission_at_sale,
            "entry_date":        entry.entry_date,
            "campaign_status":   camp.get_status_display(),
            "notes":             entry.notes or "—",
        })
    return rows


def get_postpaid_sales_summary(qs):
    """
    Aggregate totals from the PostpaidSaleEntry queryset.
    Uses frozen snapshot columns exclusively.
    """
    agg = qs.aggregate(
        total_qty=Coalesce(Sum("quantity"), Value(0)),
        total_value=Coalesce(Sum("value_at_sale"), Value(Decimal("0")), output_field=DecimalField()),
        total_commission=Coalesce(Sum("commission_at_sale"), Value(Decimal("0")), output_field=DecimalField()),
    )
    return {
        "total_entries":    qs.count(),
        "total_quantity":   agg["total_qty"],
        "total_value":      agg["total_value"],
        "total_commission": agg["total_commission"],
    }


def get_doctor_wise_totals(qs):
    """Aggregate PostpaidSaleEntry rows grouped by doctor."""
    return (
        qs.values("campaign__doctor__name")
        .annotate(
            total_qty=Coalesce(Sum("quantity"), Value(0)),
            total_value=Coalesce(Sum("value_at_sale"), Value(Decimal("0")), output_field=DecimalField()),
            total_commission=Coalesce(Sum("commission_at_sale"), Value(Decimal("0")), output_field=DecimalField()),
        )
        .order_by("-total_value")
    )


def get_rep_wise_totals(qs):
    """Aggregate PostpaidSaleEntry rows grouped by rep."""
    return (
        qs.values("rep__first_name", "rep__last_name", "rep__username")
        .annotate(
            total_qty=Coalesce(Sum("quantity"), Value(0)),
            total_value=Coalesce(Sum("value_at_sale"), Value(Decimal("0")), output_field=DecimalField()),
            total_commission=Coalesce(Sum("commission_at_sale"), Value(Decimal("0")), output_field=DecimalField()),
        )
        .order_by("-total_value")
    )


def get_medicine_wise_totals(qs):
    """Aggregate PostpaidSaleEntry rows grouped by medicine."""
    return (
        qs.values("medicine__name", "medicine__brand")
        .annotate(
            total_qty=Coalesce(Sum("quantity"), Value(0)),
            total_value=Coalesce(Sum("value_at_sale"), Value(Decimal("0")), output_field=DecimalField()),
            total_commission=Coalesce(Sum("commission_at_sale"), Value(Decimal("0")), output_field=DecimalField()),
        )
        .order_by("-total_value")
    )


def get_monthly_totals(qs):
    """Aggregate PostpaidSaleEntry rows grouped by month/year."""
    return (
        qs.values("campaign__year", "campaign__month")
        .annotate(
            total_qty=Coalesce(Sum("quantity"), Value(0)),
            total_value=Coalesce(Sum("value_at_sale"), Value(Decimal("0")), output_field=DecimalField()),
            total_commission=Coalesce(Sum("commission_at_sale"), Value(Decimal("0")), output_field=DecimalField()),
        )
        .order_by("-campaign__year", "-campaign__month")
    )


# ─────────────────────────────────────────────────────────────────────────────
# D. Settlement Ledger Report — PostpaidCampaign level
# ─────────────────────────────────────────────────────────────────────────────

def get_settlement_ledger_queryset(
    *,
    doctor_id=None,
    rep_id=None,
    month=None,
    year=None,
    status=None,
    search=None,
):
    """
    Return a filtered PostpaidCampaign queryset for the Settlement Ledger.
    Includes select_related for doctor and prefetch for payments.
    """
    qs = (
        PostpaidCampaign.objects
        .select_related("doctor", "settled_by")
        .prefetch_related("payments")
        .order_by("-year", "-month", "doctor__name")
    )

    if doctor_id:
        qs = qs.filter(doctor_id=doctor_id)
    if month:
        qs = qs.filter(month=month)
    if year:
        qs = qs.filter(year=year)
    if status:
        qs = qs.filter(status=status)
    if search:
        qs = qs.filter(
            Q(doctor__name__icontains=search) |
            Q(doctor__hospital__icontains=search)
        )
    # rep filter: campaigns whose sales entries include this rep
    if rep_id:
        campaign_ids = (
            PostpaidSaleEntry.objects
            .filter(rep_id=rep_id)
            .values_list("campaign_id", flat=True)
            .distinct()
        )
        qs = qs.filter(id__in=campaign_ids)

    return qs


def get_settlement_ledger_report(qs):
    """
    Convert a PostpaidCampaign queryset to a list of template-ready dicts
    for the Settlement Ledger Report.
    """
    rows = []
    for idx, camp in enumerate(qs.iterator(chunk_size=200), start=1):
        settled_by = camp.settled_by
        rows.append({
            "sl_no":             idx,
            "period":            f"{camp.month:02d}/{camp.year}",
            "month":             camp.month,
            "year":              camp.year,
            "doctor_name":       camp.doctor.name,
            "doctor_location":   getattr(camp.doctor, "location", "") or "—",
            "commission_pct":    camp.commission_percentage,
            "total_sales_value": camp.total_sales_value,
            "total_commission":  camp.total_commission,
            "paid_amount":       camp.paid_amount,
            "outstanding_balance": camp.outstanding_balance,
            "status":            camp.status,
            "status_display":    camp.get_status_display(),
            "is_settled":        camp.status in (PostpaidCampaign.STATUS_SETTLED, PostpaidCampaign.STATUS_LOCKED),
            "is_locked":         camp.status == PostpaidCampaign.STATUS_LOCKED,
            "settled_at":        camp.settled_at,
            "settled_by_name":   (settled_by.get_full_name() or settled_by.username) if settled_by else "—",
            "settlement_reason": camp.get_settlement_reason_display() if camp.settlement_reason else "—",
            "locked_at":         camp.locked_at,
            "campaign_id":       camp.id,
        })
    return rows


def get_settlement_summary(qs):
    """
    Aggregate totals from the PostpaidCampaign queryset for the Settlement Ledger.
    """
    agg = qs.aggregate(
        total_sales=Coalesce(Sum("total_sales_value"), Value(Decimal("0")), output_field=DecimalField()),
        total_commission=Coalesce(Sum("total_commission"), Value(Decimal("0")), output_field=DecimalField()),
        total_paid=Coalesce(Sum("paid_amount"), Value(Decimal("0")), output_field=DecimalField()),
    )
    total_commission  = agg["total_commission"]
    total_paid        = agg["total_paid"]
    total_outstanding = total_commission - total_paid

    settled_count = qs.filter(status=PostpaidCampaign.STATUS_SETTLED).count()
    locked_count  = qs.filter(status=PostpaidCampaign.STATUS_LOCKED).count()
    open_count    = qs.filter(status=PostpaidCampaign.STATUS_OPEN).count()
    partial_count = qs.filter(status=PostpaidCampaign.STATUS_PARTIAL).count()
    awaiting_count = qs.filter(status=PostpaidCampaign.STATUS_AWAITING_COMMISSION).count()

    return {
        "total_campaigns":        qs.count(),
        "total_sales":            agg["total_sales"],
        "total_commission":       total_commission,
        "total_paid":             total_paid,
        "total_outstanding":      total_outstanding,
        "settled_count":          settled_count,
        "locked_count":           locked_count,
        "open_count":             open_count,
        "partial_count":          partial_count,
        "awaiting_count":         awaiting_count,
    }


# ─────────────────────────────────────────────────────────────────────────────
# E. Excel export — Postpaid Sales Report
# ─────────────────────────────────────────────────────────────────────────────

_SALES_COLS = [
    ("Sl No",           7),
    ("Period",         10),
    ("Doctor",         24),
    ("Area",           16),
    ("Sales Rep",      20),
    ("Medicine",       24),
    ("Qty",             8),
    ("PTS at Sale (₹)", 14),
    ("Value (₹)",      14),
    ("Comm %",          9),
    ("Commission (₹)", 14),
    ("Entry Date",     14),
    ("Campaign Status", 18),
    ("Notes",          30),
]

_DOCTOR_COLS  = [("Doctor",     24), ("Qty",  8), ("Value (₹)", 14), ("Commission (₹)", 16)]
_REP_COLS     = [("Sales Rep",  24), ("Qty",  8), ("Value (₹)", 14), ("Commission (₹)", 16)]
_MEDICINE_COLS= [("Medicine",   24), ("Qty",  8), ("Value (₹)", 14), ("Commission (₹)", 16)]
_MONTHLY_COLS = [("Period",     12), ("Qty",  8), ("Value (₹)", 14), ("Commission (₹)", 16)]


def export_postpaid_sales_to_excel(sales_qs, summary):
    """
    Generate a styled multi-sheet .xlsx for the Postpaid Sales Report.

    Sheets:
      1. Sales Detail  — one row per PostpaidSaleEntry
      2. By Doctor     — doctor-wise totals
      3. By Rep        — rep-wise totals
      4. By Medicine   — medicine-wise totals
      5. By Month      — monthly totals
    """
    wb = Workbook()

    # ── Sheet 1: Sales Detail ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Sales Detail"

    ws1.merge_cells(f"A1:{get_column_letter(len(_SALES_COLS))}1")
    title = ws1["A1"]
    title.value     = "Medburg CRM — Postpaid Sales Report"
    title.font      = Font(name="Calibri", bold=True, size=14, color="1E293B")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    row = _apply_header_row(ws1, _SALES_COLS, 3)

    c_align  = Alignment(horizontal="center", vertical="center")
    r_align  = Alignment(horizontal="right",  vertical="center")
    d_align  = Alignment(vertical="center")

    for entry_row in summary.get("_rows", []):
        ws1.cell(row=row, column=1,  value=entry_row["sl_no"]).alignment = c_align
        ws1.cell(row=row, column=2,  value=entry_row["period"]).alignment = c_align
        ws1.cell(row=row, column=3,  value=entry_row["doctor_name"]).alignment = d_align
        ws1.cell(row=row, column=4,  value=entry_row["doctor_location"]).alignment = d_align
        ws1.cell(row=row, column=5,  value=entry_row["rep_name"]).alignment = d_align
        ws1.cell(row=row, column=6,  value=entry_row["medicine_name"]).alignment = d_align
        ws1.cell(row=row, column=7,  value=entry_row["quantity"]).alignment = c_align

        pts_cell = ws1.cell(row=row, column=8, value=float(entry_row["pts_at_sale"]))
        pts_cell.number_format = _CURRENCY_FMT; pts_cell.alignment = r_align

        val_cell = ws1.cell(row=row, column=9, value=float(entry_row["value_at_sale"]))
        val_cell.number_format = _CURRENCY_FMT; val_cell.alignment = r_align

        ws1.cell(row=row, column=10, value=float(entry_row["commission_pct"])).alignment = c_align

        comm_cell = ws1.cell(row=row, column=11, value=float(entry_row["commission_earned"]))
        comm_cell.number_format = _CURRENCY_FMT; comm_cell.alignment = r_align

        ws1.cell(row=row, column=12, value=entry_row["entry_date"]).alignment = c_align
        ws1.cell(row=row, column=13, value=entry_row["campaign_status"]).alignment = c_align
        ws1.cell(row=row, column=14, value=entry_row["notes"]).alignment = d_align

        if row % 2 == 0:
            for c in range(1, len(_SALES_COLS) + 1):
                ws1.cell(row=row, column=c).fill = _ALT_ROW_FILL
        for c in range(1, len(_SALES_COLS) + 1):
            ws1.cell(row=row, column=c).border = _THIN_BORDER
        row += 1

    ws1.freeze_panes = "A4"

    # Summary block on Sheet 1
    _write_summary_block(ws1, [
        ("Total Entries",        summary["total_entries"],    False),
        ("Total Quantity",       summary["total_quantity"],   False),
        ("Total Sales Value (₹)", summary["total_value"],     True),
        ("Total Commission (₹)", summary["total_commission"], True),
    ], row + 2)

    # ── Sheet 2: By Doctor ───────────────────────────────────────
    ws2 = wb.create_sheet("By Doctor")
    ws2["A1"].value = "Sales by Doctor"
    ws2["A1"].font  = Font(name="Calibri", bold=True, size=13, color="1E293B")
    dr = _apply_header_row(ws2, _DOCTOR_COLS, 3)
    for rec in summary.get("_doctor_totals", []):
        ws2.cell(row=dr, column=1, value=rec["campaign__doctor__name"])
        ws2.cell(row=dr, column=2, value=rec["total_qty"]).alignment = Alignment(horizontal="center")
        c3 = ws2.cell(row=dr, column=3, value=float(rec["total_value"]))
        c3.number_format = _CURRENCY_FMT
        c4 = ws2.cell(row=dr, column=4, value=float(rec["total_commission"]))
        c4.number_format = _CURRENCY_FMT
        for c in range(1, 5): ws2.cell(row=dr, column=c).border = _THIN_BORDER
        dr += 1

    # ── Sheet 3: By Rep ──────────────────────────────────────────
    ws3 = wb.create_sheet("By Rep")
    ws3["A1"].value = "Sales by Rep"
    ws3["A1"].font  = Font(name="Calibri", bold=True, size=13, color="1E293B")
    rr = _apply_header_row(ws3, _REP_COLS, 3)
    for rec in summary.get("_rep_totals", []):
        fn = rec.get("rep__first_name") or ""
        ln = rec.get("rep__last_name")  or ""
        un = rec.get("rep__username")   or ""
        name = (f"{fn} {ln}".strip()) or un
        ws3.cell(row=rr, column=1, value=name)
        ws3.cell(row=rr, column=2, value=rec["total_qty"]).alignment = Alignment(horizontal="center")
        c3 = ws3.cell(row=rr, column=3, value=float(rec["total_value"]))
        c3.number_format = _CURRENCY_FMT
        c4 = ws3.cell(row=rr, column=4, value=float(rec["total_commission"]))
        c4.number_format = _CURRENCY_FMT
        for c in range(1, 5): ws3.cell(row=rr, column=c).border = _THIN_BORDER
        rr += 1

    # ── Sheet 4: By Medicine ─────────────────────────────────────
    ws4 = wb.create_sheet("By Medicine")
    ws4["A1"].value = "Sales by Medicine"
    ws4["A1"].font  = Font(name="Calibri", bold=True, size=13, color="1E293B")
    mr = _apply_header_row(ws4, _MEDICINE_COLS, 3)
    for rec in summary.get("_medicine_totals", []):
        med_name  = rec.get("medicine__name") or "—"
        med_brand = rec.get("medicine__brand") or ""
        label = f"{med_name} ({med_brand})" if med_brand else med_name
        ws4.cell(row=mr, column=1, value=label)
        ws4.cell(row=mr, column=2, value=rec["total_qty"]).alignment = Alignment(horizontal="center")
        c3 = ws4.cell(row=mr, column=3, value=float(rec["total_value"]))
        c3.number_format = _CURRENCY_FMT
        c4 = ws4.cell(row=mr, column=4, value=float(rec["total_commission"]))
        c4.number_format = _CURRENCY_FMT
        for c in range(1, 5): ws4.cell(row=mr, column=c).border = _THIN_BORDER
        mr += 1

    # ── Sheet 5: By Month ────────────────────────────────────────
    ws5 = wb.create_sheet("By Month")
    ws5["A1"].value = "Sales by Month"
    ws5["A1"].font  = Font(name="Calibri", bold=True, size=13, color="1E293B")
    mo = _apply_header_row(ws5, _MONTHLY_COLS, 3)
    for rec in summary.get("_monthly_totals", []):
        period = f"{rec['campaign__month']:02d}/{rec['campaign__year']}"
        ws5.cell(row=mo, column=1, value=period).alignment = Alignment(horizontal="center")
        ws5.cell(row=mo, column=2, value=rec["total_qty"]).alignment = Alignment(horizontal="center")
        c3 = ws5.cell(row=mo, column=3, value=float(rec["total_value"]))
        c3.number_format = _CURRENCY_FMT
        c4 = ws5.cell(row=mo, column=4, value=float(rec["total_commission"]))
        c4.number_format = _CURRENCY_FMT
        for c in range(1, 5): ws5.cell(row=mo, column=c).border = _THIN_BORDER
        mo += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# F. Excel export — Settlement Ledger Report
# ─────────────────────────────────────────────────────────────────────────────

_LEDGER_COLS = [
    ("Sl No",            7),
    ("Period",          10),
    ("Doctor",          24),
    ("Area",            16),
    ("Comm %",           9),
    ("Sales Value (₹)", 16),
    ("Commission (₹)",  16),
    ("Paid (₹)",        14),
    ("Outstanding (₹)", 16),
    ("Status",          18),
    ("Settlement Reason", 22),
    ("Settled At",      16),
    ("Settled By",      20),
    ("Locked At",       16),
]


def export_settlement_ledger_to_excel(ledger_rows, summary):
    """
    Generate a styled .xlsx for the Settlement Ledger Report.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Settlement Ledger"

    ws.merge_cells(f"A1:{get_column_letter(len(_LEDGER_COLS))}1")
    title = ws["A1"]
    title.value     = "Medburg CRM — Postpaid Settlement Ledger"
    title.font      = Font(name="Calibri", bold=True, size=14, color="1E293B")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    row = _apply_header_row(ws, _LEDGER_COLS, 3)

    c_align = Alignment(horizontal="center", vertical="center")
    r_align = Alignment(horizontal="right",  vertical="center")
    d_align = Alignment(vertical="center")

    for lr in ledger_rows:
        ws.cell(row=row, column=1,  value=lr["sl_no"]).alignment = c_align
        ws.cell(row=row, column=2,  value=lr["period"]).alignment = c_align
        ws.cell(row=row, column=3,  value=lr["doctor_name"]).alignment = d_align
        ws.cell(row=row, column=4,  value=lr["doctor_location"]).alignment = d_align
        ws.cell(row=row, column=5,  value=float(lr["commission_pct"]) if lr["commission_pct"] is not None else "—").alignment = c_align

        vs = ws.cell(row=row, column=6,  value=float(lr["total_sales_value"]))
        vs.number_format = _CURRENCY_FMT; vs.alignment = r_align

        vc = ws.cell(row=row, column=7,  value=float(lr["total_commission"]))
        vc.number_format = _CURRENCY_FMT; vc.alignment = r_align

        vp = ws.cell(row=row, column=8,  value=float(lr["paid_amount"]))
        vp.number_format = _CURRENCY_FMT; vp.alignment = r_align

        vo = ws.cell(row=row, column=9,  value=float(lr["outstanding_balance"]))
        vo.number_format = _CURRENCY_FMT; vo.alignment = r_align

        ws.cell(row=row, column=10, value=lr["status_display"]).alignment = c_align
        ws.cell(row=row, column=11, value=lr["settlement_reason"]).alignment = d_align
        ws.cell(row=row, column=12, value=lr["settled_at"]).alignment = c_align
        ws.cell(row=row, column=13, value=lr["settled_by_name"]).alignment = d_align
        ws.cell(row=row, column=14, value=lr["locked_at"]).alignment = c_align

        if row % 2 == 0:
            for c in range(1, len(_LEDGER_COLS) + 1):
                ws.cell(row=row, column=c).fill = _ALT_ROW_FILL
        for c in range(1, len(_LEDGER_COLS) + 1):
            ws.cell(row=row, column=c).border = _THIN_BORDER
        row += 1

    ws.freeze_panes = "A4"

    _write_summary_block(ws, [
        ("Total Campaigns",       summary["total_campaigns"],  False),
        ("Total Sales Value (₹)", summary["total_sales"],      True),
        ("Total Commission (₹)",  summary["total_commission"], True),
        ("Total Paid (₹)",        summary["total_paid"],       True),
        ("Total Outstanding (₹)", summary["total_outstanding"], True),
        ("Settled Campaigns",     summary["settled_count"],    False),
        ("Locked Campaigns",      summary["locked_count"],     False),
        ("Partial Campaigns",     summary["partial_count"],    False),
        ("Open Campaigns",        summary["open_count"],       False),
        ("Awaiting Commission",   summary["awaiting_count"],   False),
    ], row + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
