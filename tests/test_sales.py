from decimal import Decimal
from datetime import date, timedelta
from django.test import TestCase
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.contrib.auth import get_user_model

from doctors.models import Doctor, DoctorMedicine, Investment
from medicines.models import Medicine
from sales.models import SalesEntry, PostpaidCampaign, PostpaidSaleEntry, CampaignPayment, PostpaidCampaignCorrection
from doctors.services.doctor_service import (
    get_dashboard_alerts,
    get_dashboard_queryset,
    get_prepaid_admin_metrics,
    get_unified_activity_feed,
    get_active_postpaid_campaigns,
)
from sales.services.postpaid_service import (
    get_postpaid_sales_queryset,
    get_postpaid_sales_report,
    get_postpaid_sales_summary,
    get_settlement_ledger_queryset,
    get_settlement_ledger_report,
    get_settlement_summary,
    export_postpaid_sales_to_excel,
    export_settlement_ledger_to_excel,
)

User = get_user_model()


class PostpaidLifecycleTests(TestCase):
    def setUp(self):
        self.rep = User.objects.create_user(username="rep_user", password="pwd", role="rep")
        self.admin = User.objects.create_user(username="admin_user", password="pwd", role="admin")

        self.doctor = Doctor.objects.create(
            name="Dr. Postpaid",
            mode="postpaid",
            assigned_rep=self.rep,
            is_active=True
        )
        self.medicine = Medicine.objects.create(
            name="Med A",
            pts=Decimal("150.00"),
            ptr=Decimal("130.00"),
            mrp=Decimal("180.00"),
            is_active=True
        )
        DoctorMedicine.objects.create(doctor=self.doctor, medicine=self.medicine)

    def test_postpaid_sale_entry_auto_creates_campaign_awaiting_commission(self):
        """
        Submitting a postpaid sales entry should auto-create a PostpaidCampaign
        in 'awaiting_commission' status with a commission_percentage of None.
        """
        # Create campaign and entry via model creation
        campaign, created = PostpaidCampaign.objects.get_or_create(
            doctor=self.doctor,
            month=6,
            year=2026,
            defaults={
                "commission_percentage": None,
                "status": PostpaidCampaign.STATUS_AWAITING_COMMISSION,
            }
        )
        self.assertTrue(created)
        self.assertEqual(campaign.status, PostpaidCampaign.STATUS_AWAITING_COMMISSION)
        self.assertIsNone(campaign.commission_percentage)

        sale = PostpaidSaleEntry.objects.create(
            campaign=campaign,
            medicine=self.medicine,
            quantity=10,
            entry_date=date.today(),
            rep=self.rep
        )

        # Check snapshots
        self.assertEqual(sale.pts_at_sale, Decimal("150.00"))
        self.assertEqual(sale.value_at_sale, Decimal("1500.00"))
        self.assertEqual(sale.commission_percentage_at_sale, Decimal("0.00"))
        self.assertEqual(sale.commission_at_sale, Decimal("0.00"))

        # Verify campaign totals recalculate
        campaign.refresh_from_db()
        self.assertEqual(campaign.total_sales_value, Decimal("1500.00"))
        self.assertEqual(campaign.total_commission, Decimal("0.00"))

    def test_commission_update_transitions_status_to_open(self):
        """
        Setting a commission percentage on an awaiting_commission campaign
        should automatically transition the campaign to Open status.
        """
        campaign = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=6,
            year=2026,
            commission_percentage=None,
            status=PostpaidCampaign.STATUS_AWAITING_COMMISSION
        )
        self.assertEqual(campaign.status, PostpaidCampaign.STATUS_AWAITING_COMMISSION)

        # Configure commission
        campaign.update_commission_percentage(Decimal("10.00"))
        self.assertEqual(campaign.status, PostpaidCampaign.STATUS_OPEN)
        self.assertEqual(campaign.commission_percentage, Decimal("10.00"))

    def test_postpaid_snapshot_rule(self):
        """
        PostpaidSaleEntry values are frozen at creation and never recalculate
        from Medicine.pts even if the medicine price changes later.
        """
        campaign = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=6,
            year=2026,
            commission_percentage=Decimal("15.00"),
            status=PostpaidCampaign.STATUS_OPEN
        )

        sale = PostpaidSaleEntry.objects.create(
            campaign=campaign,
            medicine=self.medicine,
            quantity=10,
            entry_date=date.today(),
            rep=self.rep
        )

        self.assertEqual(sale.pts_at_sale, Decimal("150.00"))
        self.assertEqual(sale.value_at_sale, Decimal("1500.00"))
        self.assertEqual(sale.commission_at_sale, Decimal("225.00"))

        # Alter live medicine price
        self.medicine.pts = Decimal("200.00")
        self.medicine.save()

        # Save sale entry again
        sale.save()

        # Verify snapshot prices DID NOT change
        sale.refresh_from_db()
        self.assertEqual(sale.pts_at_sale, Decimal("150.00"))
        self.assertEqual(sale.value_at_sale, Decimal("1500.00"))
        self.assertEqual(sale.commission_at_sale, Decimal("225.00"))

    def test_payments_only_allowed_on_partial_status(self):
        """
        Payments cannot be recorded unless the campaign is in Partial status.
        This blocks payments on Awaiting Commission, Open, Settled, and Locked campaigns.
        """
        campaign = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=6,
            year=2026,
            commission_percentage=None,
            status=PostpaidCampaign.STATUS_AWAITING_COMMISSION
        )

        # Awaiting Commission status
        payment = CampaignPayment(campaign=campaign, amount=Decimal("100.00"), payment_date=date.today())
        with self.assertRaises(ValidationError):
            payment.clean()

        # Open status
        campaign.update_commission_percentage(Decimal("10.00"))
        self.assertEqual(campaign.status, PostpaidCampaign.STATUS_OPEN)
        with self.assertRaises(ValidationError):
            payment.clean()

        # Record sales so there is commission target (must be done while status is Open)
        PostpaidSaleEntry.objects.create(
            campaign=campaign,
            medicine=self.medicine,
            quantity=10,
            entry_date=date.today(),
            rep=self.rep
        )
        campaign.refresh_from_db()
        self.assertEqual(campaign.total_commission, Decimal("150.00"))

        # Advance manually to Partial status
        campaign.status = PostpaidCampaign.STATUS_PARTIAL
        campaign.save()

        # Try now — should pass validation
        payment.clean()  # Should not raise ValidationError

        # Save payment (transitions to Settled automatically if fully paid)
        payment.save()
        campaign.refresh_from_db()
        self.assertEqual(campaign.status, PostpaidCampaign.STATUS_PARTIAL)  # 100 < 150

        # Create second payment to meet balance (should NOT auto-settle campaign)
        payment2 = CampaignPayment.objects.create(
            campaign=campaign,
            amount=Decimal("50.00"),
            payment_date=date.today()
        )
        campaign.refresh_from_db()
        # Verify that it does NOT auto-settle anymore
        self.assertEqual(campaign.status, PostpaidCampaign.STATUS_PARTIAL)

        # Manually settle the campaign (allowed since outstanding balance is 0.00)
        campaign.status = PostpaidCampaign.STATUS_SETTLED
        campaign.save()

        # Record a payment when campaign is Settled (should raise ValidationError)
        payment3 = CampaignPayment(campaign=campaign, amount=Decimal("10.00"), payment_date=date.today())
        with self.assertRaises(ValidationError):
            payment3.clean()

    def test_locked_campaigns_block_all_modifications(self):
        """
        Locked campaigns block all edits to campaign details.
        """
        campaign = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=6,
            year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_LOCKED
        )

        # Attempt to change commission percentage
        campaign.commission_percentage = Decimal("15.00")
        with self.assertRaises(ValidationError):
            campaign.clean()

    def test_commission_locked_past_open(self):
        """
        Commission percentage cannot be changed once the campaign transitions past Open.
        """
        campaign = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=6,
            year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_PARTIAL
        )

        # Attempt to modify commission
        campaign.commission_percentage = Decimal("15.00")
        with self.assertRaises(ValidationError):
            campaign.clean()

    def test_append_only_ledger_entries(self):
        """
        Ledger entries cannot be updated or deleted.
        """
        campaign = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=6,
            year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_PARTIAL
        )

        payment = CampaignPayment.objects.create(
            campaign=campaign,
            amount=Decimal("100.00"),
            payment_date=date.today()
        )

        # Attempt to edit payment amount
        payment.amount = Decimal("150.00")
        with self.assertRaises(ValueError):
            payment.save()

        # Attempt to delete payment
        with self.assertRaises(ValidationError):
            payment.delete()

    def test_awaiting_commission_alerts(self):
        """
        Unconfigured campaigns show warnings after 3 days (yellow) and 7 days (red).
        """
        # Clear existing campaigns to avoid alert pollution
        PostpaidCampaign.objects.all().delete()

        # Create one awaiting commission campaign with age = 4 days
        campaign_yellow = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=5,
            year=2026,
            commission_percentage=None,
            status=PostpaidCampaign.STATUS_AWAITING_COMMISSION
        )
        # Manually alter created_at to simulate age
        PostpaidCampaign.objects.filter(pk=campaign_yellow.pk).update(
            created_at=timezone.now() - timedelta(days=4)
        )

        # Create another awaiting commission campaign with age = 8 days
        campaign_red = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=4,
            year=2026,
            commission_percentage=None,
            status=PostpaidCampaign.STATUS_AWAITING_COMMISSION
        )
        PostpaidCampaign.objects.filter(pk=campaign_red.pk).update(
            created_at=timezone.now() - timedelta(days=8)
        )

        alerts = get_dashboard_alerts()

        # Look for warnings in alerts list
        yellow_found = any("WARNING" in a["message"] and "Dr. Postpaid" in a["message"] and a["type"] == "warning" for a in alerts)
        red_found = any("RED WARNING" in a["message"] and "Dr. Postpaid" in a["message"] and a["type"] == "danger" for a in alerts)

        self.assertTrue(yellow_found)
        self.assertTrue(red_found)


class PrepaidDashboardTests(TestCase):
    def setUp(self):
        self.rep = get_user_model().objects.create_user(username="rep_user2", password="pwd", role="rep")
        self.doctor = Doctor.objects.create(
            name="Dr. Prepaid ABC",
            mode="prepaid",
            assigned_rep=self.rep,
            is_active=True
        )
        self.medicine = Medicine.objects.create(
            name="Med B",
            pts=Decimal("100.00"),
            ptr=Decimal("80.00"),
            mrp=Decimal("120.00"),
            is_active=True
        )
        DoctorMedicine.objects.create(doctor=self.doctor, medicine=self.medicine)

    def test_dashboard_excludes_completed_investments_from_active_exposure(self):
        """
        Prepaid dashboard queryset achieved_roi must only sum sales entries from active (in_progress) investments.
        Completed investment totals and sales must not contaminate active dashboard metrics.
        """
        # Create completed investment
        inv_completed = Investment.objects.create(
            doctor=self.doctor,
            amount=Decimal("1000.00"),
            roi_ratio=Decimal("2.50"),
            start_date=date.today() - timedelta(days=60),
            status=Investment.STATUS_IN_PROGRESS
        )
        
        # Create sales entry associated with completed investment
        entry_completed = SalesEntry.objects.create(
            rep=self.rep,
            doctor=self.doctor,
            investment=inv_completed,
            medicine=self.medicine,
            quantity=25,  # 25 * 100 = 2500 achieved (fully completes target of 1000*2.5 = 2500)
            entry_date=date.today() - timedelta(days=50)
        )
        inv_completed.refresh_from_db()
        self.assertEqual(inv_completed.status, Investment.STATUS_IN_PROGRESS)
        # Admin manually sets status to completed (allowed since balance <= 0)
        inv_completed.status = Investment.STATUS_COMPLETED
        inv_completed.save()
        self.assertEqual(inv_completed.status, Investment.STATUS_COMPLETED)
        # Ensure snapshot populated correctly
        self.assertEqual(entry_completed.value_at_sale, Decimal("2500.00"))

        # Verify that doctor with only completed investments returns 0 metrics on the active dashboard
        qs1 = get_dashboard_queryset().filter(pk=self.doctor.pk)
        self.assertEqual(qs1.count(), 1)
        doc_metric1 = qs1.first()
        self.assertEqual(doc_metric1.total_investment, Decimal("0.00"))
        self.assertEqual(doc_metric1.total_roi_amount, Decimal("0.00"))
        self.assertEqual(doc_metric1.achieved_roi, Decimal("0.00"))
        self.assertEqual(doc_metric1.balance_roi, Decimal("0.00"))

        # Now create an active (in_progress) investment
        inv_active = Investment.objects.create(
            doctor=self.doctor,
            amount=Decimal("2000.00"),
            roi_ratio=Decimal("2.00"),
            start_date=date.today() - timedelta(days=10),
            status=Investment.STATUS_IN_PROGRESS
        )

        # Create sales entry associated with active investment
        entry_active = SalesEntry.objects.create(
            rep=self.rep,
            doctor=self.doctor,
            investment=inv_active,
            medicine=self.medicine,
            quantity=15,  # 15 * 100 = 1500 achieved
            entry_date=date.today()
        )

        # Verify that the dashboard queryset now only aggregates metrics for the active investment
        qs2 = get_dashboard_queryset().filter(pk=self.doctor.pk)
        doc_metric2 = qs2.first()
        self.assertEqual(doc_metric2.total_investment, Decimal("2000.00"))
        self.assertEqual(doc_metric2.total_roi_amount, Decimal("4000.00"))
        self.assertEqual(doc_metric2.achieved_roi, Decimal("1500.00"))  # excludes the 2500 from inv_completed
        self.assertEqual(doc_metric2.balance_roi, Decimal("2500.00"))   # 4000 - 1500 = 2500


class DashboardStabilizationTests(TestCase):
    def setUp(self):
        self.rep = User.objects.create_user(username="rep_dashboard", password="pwd", role="rep")
        self.medicine = Medicine.objects.create(
            name="Test Med",
            pts=Decimal("100.00"),
            ptr=Decimal("80.00"),
            mrp=Decimal("120.00"),
            is_active=True
        )

        self.prepaid_doc = Doctor.objects.create(
            name="Dr. Prepaid Dashboard",
            mode="prepaid",
            assigned_rep=self.rep,
            is_active=True
        )
        DoctorMedicine.objects.create(doctor=self.prepaid_doc, medicine=self.medicine)

        self.postpaid_doc = Doctor.objects.create(
            name="Dr. Postpaid Dashboard",
            mode="postpaid",
            assigned_rep=self.rep,
            is_active=True
        )
        DoctorMedicine.objects.create(doctor=self.postpaid_doc, medicine=self.medicine)

    def test_recovery_rate_active_records_only_and_not_clamped(self):
        """
        Validate that active recovery rate works only on STATUS_IN_PROGRESS investments
        and does not clamp at 100% (can exceed 100%).
        """
        # Completed investment (should be ignored by active recovery rate)
        inv_comp = Investment.objects.create(
            doctor=self.prepaid_doc,
            amount=Decimal("1000.00"),
            roi_ratio=Decimal("2.00"),
            start_date=date.today() - timedelta(days=60),
            status=Investment.STATUS_IN_PROGRESS
        )
        SalesEntry.objects.create(
            rep=self.rep,
            doctor=self.prepaid_doc,
            investment=inv_comp,
            medicine=self.medicine,
            quantity=20,
            entry_date=date.today() - timedelta(days=50)
        )
        inv_comp.status = Investment.STATUS_COMPLETED
        inv_comp.save()
        inv_comp.refresh_from_db()
        self.assertEqual(inv_comp.status, Investment.STATUS_COMPLETED)

        # Active investment with over-recovery (3000 achieved on 2000 target = 150%)
        inv_active = Investment.objects.create(
            doctor=self.prepaid_doc,
            amount=Decimal("1000.00"),
            roi_ratio=Decimal("2.00"),
            start_date=date.today() - timedelta(days=10),
            status=Investment.STATUS_IN_PROGRESS
        )
        SalesEntry.objects.create(
            rep=self.rep,
            doctor=self.prepaid_doc,
            investment=inv_active,
            medicine=self.medicine,
            quantity=30,
            entry_date=date.today()
        )
        inv_active.refresh_from_db()
        self.assertEqual(inv_active.status, Investment.STATUS_IN_PROGRESS)

        metrics = get_prepaid_admin_metrics()
        self.assertAlmostEqual(metrics["recovery_rate"], Decimal("150.00"))

    def test_settlement_integrity_alert(self):
        """
        Verify that a settled campaign with a computed outstanding balance > 0
        triggers a danger-type alert in the dashboard.
        """
        # Clear existing campaigns to avoid alert pollution
        PostpaidCampaign.objects.all().delete()

        camp = PostpaidCampaign.objects.create(
            doctor=self.postpaid_doc,
            month=6,
            year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_SETTLED,
            total_sales_value=Decimal("10000.00"),
            total_commission=Decimal("1000.00"),
            paid_amount=Decimal("800.00"),  # Outstanding = 200 > 0
            settlement_reason=PostpaidCampaign.REASON_WRITE_OFF,
            settlement_notes="Approved write-off for deviation."
        )

        alerts = get_dashboard_alerts()
        integrity_alerts = [a for a in alerts if a["type"] == "danger" and "Settlement Integrity Anomaly" in a["message"]]
        self.assertEqual(len(integrity_alerts), 1)
        self.assertIn("Dr. Postpaid Dashboard", integrity_alerts[0]["message"])
        self.assertIn("₹200.00", integrity_alerts[0]["message"])

    def test_completed_investments_rolling_30_days(self):
        """
        Verify completed investments count correctly filters based on a rolling 30-day window.
        """
        # Clean existing investments
        Investment.objects.all().delete()

        # Completed within 30 days
        inv1 = Investment.objects.create(
            doctor=self.prepaid_doc,
            amount=Decimal("1000.00"),
            roi_ratio=Decimal("2.00"),
            start_date=date.today() - timedelta(days=20),
            status=Investment.STATUS_COMPLETED
        )
        Investment.objects.filter(pk=inv1.pk).update(updated_at=timezone.now() - timedelta(days=10))

        # Completed more than 30 days ago
        inv2 = Investment.objects.create(
            doctor=self.prepaid_doc,
            amount=Decimal("1000.00"),
            roi_ratio=Decimal("2.00"),
            start_date=date.today() - timedelta(days=50),
            status=Investment.STATUS_COMPLETED
        )
        Investment.objects.filter(pk=inv2.pk).update(updated_at=timezone.now() - timedelta(days=40))

        # Active investment (not completed)
        inv3 = Investment.objects.create(
            doctor=self.prepaid_doc,
            amount=Decimal("1000.00"),
            roi_ratio=Decimal("2.00"),
            start_date=date.today() - timedelta(days=5),
            status=Investment.STATUS_IN_PROGRESS
        )
        Investment.objects.filter(pk=inv3.pk).update(updated_at=timezone.now() - timedelta(days=2))

        metrics = get_prepaid_admin_metrics()
        self.assertEqual(metrics["completed_last_30d"], 1)

    def test_activity_feed_merge_and_ordering(self):
        """
        Verify that the activity feed correctly interleaves prepaid and postpaid sales
        in descending chronological order (entry_date, created_at).
        """
        # Clean existing entries
        SalesEntry.objects.all().delete()
        PostpaidSaleEntry.objects.all().delete()
        PostpaidCampaign.objects.all().delete()

        inv = Investment.objects.create(
            doctor=self.prepaid_doc,
            amount=Decimal("1000.00"),
            roi_ratio=Decimal("2.00"),
            start_date=date.today() - timedelta(days=10),
            status=Investment.STATUS_IN_PROGRESS
        )
        camp = PostpaidCampaign.objects.create(
            doctor=self.postpaid_doc,
            month=6,
            year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_OPEN
        )

        # 1. Old prepaid entry (2 days ago)
        se1 = SalesEntry.objects.create(
            rep=self.rep,
            doctor=self.prepaid_doc,
            investment=inv,
            medicine=self.medicine,
            quantity=10,
            entry_date=date.today() - timedelta(days=2)
        )
        SalesEntry.objects.filter(pk=se1.pk).update(created_at=timezone.now() - timedelta(days=2))

        # 2. Mid postpaid entry (1 day ago)
        pse1 = PostpaidSaleEntry.objects.create(
            campaign=camp,
            medicine=self.medicine,
            quantity=5,
            entry_date=date.today() - timedelta(days=1),
            rep=self.rep
        )
        PostpaidSaleEntry.objects.filter(pk=pse1.pk).update(created_at=timezone.now() - timedelta(days=1))

        # 3. New prepaid entry (today)
        se2 = SalesEntry.objects.create(
            rep=self.rep,
            doctor=self.prepaid_doc,
            investment=inv,
            medicine=self.medicine,
            quantity=10,
            entry_date=date.today()
        )
        SalesEntry.objects.filter(pk=se2.pk).update(created_at=timezone.now())

        feed = get_unified_activity_feed()

        # Should return newest first
        self.assertEqual(len(feed), 3)
        self.assertEqual(feed[0]["type"], "prepaid")
        self.assertEqual(feed[0]["obj"].pk, se2.pk)

        self.assertEqual(feed[1]["type"], "postpaid")
        self.assertEqual(feed[1]["obj"].pk, pse1.pk)

        self.assertEqual(feed[2]["type"], "prepaid")
        self.assertEqual(feed[2]["obj"].pk, se1.pk)

    def test_activity_feed_limit_enforcement(self):
        """
        Verify that creating 15 prepaid and 15 postpaid entries results in a feed
        limited to exactly 10 rows.
        """
        # Clean existing entries
        SalesEntry.objects.all().delete()
        PostpaidSaleEntry.objects.all().delete()
        PostpaidCampaign.objects.all().delete()

        inv = Investment.objects.create(
            doctor=self.prepaid_doc,
            amount=Decimal("1000.00"),
            roi_ratio=Decimal("2.00"),
            start_date=date.today() - timedelta(days=10),
            status=Investment.STATUS_IN_PROGRESS
        )
        camp = PostpaidCampaign.objects.create(
            doctor=self.postpaid_doc,
            month=6,
            year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_OPEN
        )

        for i in range(15):
            SalesEntry.objects.create(
                rep=self.rep,
                doctor=self.prepaid_doc,
                investment=inv,
                medicine=self.medicine,
                quantity=1,
                entry_date=date.today() - timedelta(days=i)
            )
            PostpaidSaleEntry.objects.create(
                campaign=camp,
                medicine=self.medicine,
                quantity=1,
                entry_date=date.today() - timedelta(days=i),
                rep=self.rep
            )

        feed = get_unified_activity_feed()
        self.assertEqual(len(feed), 10)


class SettlementFoundationTests(TestCase):
    def setUp(self):
        self.rep = User.objects.create_user(username="rep_user_mr4g", password="pwd", role="rep")
        self.admin = User.objects.create_user(username="admin_user_mr4g", password="pwd", role="admin")
        self.doctor = Doctor.objects.create(
            name="Dr. MR4G",
            mode="postpaid",
            assigned_rep=self.rep,
            is_active=True
        )
        self.medicine = Medicine.objects.create(
            name="Med MR4G",
            pts=Decimal("100.00"),
            ptr=Decimal("80.00"),
            mrp=Decimal("120.00"),
            is_active=True
        )
        DoctorMedicine.objects.create(doctor=self.doctor, medicine=self.medicine)

    def test_campaign_deletion_blocks(self):
        """
        Deleting a Settled or Locked campaign raises ValidationError. Deleting Open/Awaiting is allowed.
        """
        # 1. Awaiting Commission (Allowed)
        campaign_awaiting = PostpaidCampaign.objects.create(
            doctor=self.doctor, month=1, year=2026, status=PostpaidCampaign.STATUS_AWAITING_COMMISSION
        )
        campaign_awaiting.delete()  # Should not raise error

        # 2. Open (Allowed)
        campaign_open = PostpaidCampaign.objects.create(
            doctor=self.doctor, month=2, year=2026, status=PostpaidCampaign.STATUS_OPEN, commission_percentage=Decimal("10.00")
        )
        campaign_open.delete()  # Should not raise error

        # 3. Partial (Blocked)
        campaign_partial = PostpaidCampaign.objects.create(
            doctor=self.doctor, month=3, year=2026, status=PostpaidCampaign.STATUS_PARTIAL, commission_percentage=Decimal("10.00")
        )
        with self.assertRaises(ValidationError):
            campaign_partial.delete()

        # 4. Settled (Blocked)
        campaign_settled = PostpaidCampaign.objects.create(
            doctor=self.doctor, month=4, year=2026, status=PostpaidCampaign.STATUS_SETTLED, commission_percentage=Decimal("10.00"),
            settlement_reason=PostpaidCampaign.REASON_WRITE_OFF, settlement_notes="Settle it"
        )
        with self.assertRaises(ValidationError):
            campaign_settled.delete()

        # 5. Locked (Blocked)
        campaign_locked = PostpaidCampaign.objects.create(
            doctor=self.doctor, month=5, year=2026, status=PostpaidCampaign.STATUS_LOCKED, commission_percentage=Decimal("10.00")
        )
        with self.assertRaises(ValidationError):
            campaign_locked.delete()

    def test_sales_entry_deletion_blocks_and_recalculation(self):
        """
        Deleting a sales entry on Partial/Settled/Locked campaigns raises ValidationError.
        Deleting a sales entry on Open/Awaiting campaigns is allowed and triggers recalculation of campaign sales totals.
        """
        campaign = PostpaidCampaign.objects.create(
            doctor=self.doctor, month=1, year=2026, status=PostpaidCampaign.STATUS_OPEN, commission_percentage=Decimal("10.00")
        )
        sale1 = PostpaidSaleEntry.objects.create(
            campaign=campaign, medicine=self.medicine, quantity=10, entry_date=date.today(), rep=self.rep
        )
        sale2 = PostpaidSaleEntry.objects.create(
            campaign=campaign, medicine=self.medicine, quantity=5, entry_date=date.today(), rep=self.rep
        )
        campaign.refresh_from_db()
        self.assertEqual(campaign.total_sales_value, Decimal("1500.00"))
        self.assertEqual(campaign.total_commission, Decimal("150.00"))

        # Deletion on Open is allowed and triggers recalculation
        sale2.delete()
        campaign.refresh_from_db()
        self.assertEqual(campaign.total_sales_value, Decimal("1000.00"))
        self.assertEqual(campaign.total_commission, Decimal("100.00"))

        # Deletion on Partial is blocked
        campaign.status = PostpaidCampaign.STATUS_PARTIAL
        campaign.save()
        with self.assertRaises(ValidationError):
            sale1.delete()

    def test_settlement_checklist_validation(self):
        """
        Advancing a campaign with outstanding_balance > 0 to Settled status fails if settlement_reason or settlement_notes is missing, but passes if they are provided.
        Fully paid campaigns settle without reason/notes.
        """
        campaign = PostpaidCampaign.objects.create(
            doctor=self.doctor, month=1, year=2026, status=PostpaidCampaign.STATUS_OPEN, commission_percentage=Decimal("10.00")
        )
        PostpaidSaleEntry.objects.create(
            campaign=campaign, medicine=self.medicine, quantity=10, entry_date=date.today(), rep=self.rep
        )
        campaign.status = PostpaidCampaign.STATUS_PARTIAL
        campaign.save()
        campaign.refresh_status()
        self.assertEqual(campaign.outstanding_balance, Decimal("100.00"))

        # 1. Try to settle without reason/notes -> should fail
        campaign.status = PostpaidCampaign.STATUS_SETTLED
        with self.assertRaises(ValidationError) as ctx:
            campaign.full_clean()
        self.assertIn("settlement_reason", ctx.exception.message_dict)

        # 2. Try with only reason -> should fail
        campaign.settlement_reason = PostpaidCampaign.REASON_WRITE_OFF
        with self.assertRaises(ValidationError) as ctx:
            campaign.full_clean()
        self.assertIn("settlement_notes", ctx.exception.message_dict)

        # 3. Try with reason and notes -> should pass
        campaign.settlement_notes = "Approved write-off for deviation."
        campaign.full_clean()  # Should not raise error
        campaign.save()

        # 4. Try settling fully paid campaign without reason/notes -> should pass
        campaign2 = PostpaidCampaign.objects.create(
            doctor=self.doctor, month=2, year=2026, status=PostpaidCampaign.STATUS_OPEN, commission_percentage=Decimal("10.00")
        )
        PostpaidSaleEntry.objects.create(
            campaign=campaign2, medicine=self.medicine, quantity=10, entry_date=date.today(), rep=self.rep
        )
        campaign2.status = PostpaidCampaign.STATUS_PARTIAL
        campaign2.save()
        campaign2.refresh_status()
        # Record payment matching total commission (100.00)
        CampaignPayment.objects.create(campaign=campaign2, amount=Decimal("100.00"), payment_date=date.today())
        campaign2.refresh_from_db()
        self.assertEqual(campaign2.outstanding_balance, Decimal("0.00"))

        campaign2.status = PostpaidCampaign.STATUS_SETTLED
        campaign2.full_clean()  # Should not raise error since balance is 0.00
        campaign2.save()

    def test_transaction_integrity(self):
        """
        Ensure database errors roll back modifications successfully.
        """
        from django.db import transaction
        
        campaign = PostpaidCampaign.objects.create(
            doctor=self.doctor, month=1, year=2026, status=PostpaidCampaign.STATUS_OPEN, commission_percentage=Decimal("10.00")
        )
        
        try:
            with transaction.atomic():
                PostpaidSaleEntry.objects.create(
                    campaign=campaign, medicine=self.medicine, quantity=10, entry_date=date.today(), rep=self.rep
                )
                # Force an ValidationError by violating unique constraint on campaign (doctor+month+year)
                PostpaidCampaign.objects.create(
                    doctor=self.doctor, month=1, year=2026
                )
        except ValidationError:
            pass

        # Verify that the PostpaidSaleEntry was rolled back and campaign totals are 0
        campaign.refresh_from_db()
        self.assertEqual(campaign.sales_entries.count(), 0)
        self.assertEqual(campaign.total_sales_value, Decimal("0.00"))


# ─────────────────────────────────────────────────────────────────────────────
# MR-8.0: Audit & Correction Layer
# ─────────────────────────────────────────────────────────────────────────────


class AuditCorrectionLayerTests(TestCase):
    """
    Tests for MR-8.0 objectives:
    A. Legacy purge verification (PostpaidEntry gone).
    B. Ledger protection for Locked/Settled campaigns.
    C. PostpaidCampaignCorrection model correctness.
    D. Admin integration asserted via model-layer rules.
    """

    def setUp(self):
        self.rep = User.objects.create_user(
            username="mr8_rep", password="pwd", role="rep"
        )
        self.admin = User.objects.create_user(
            username="mr8_admin", password="pwd", role="admin"
        )
        self.doctor = Doctor.objects.create(
            name="Dr. MR8",
            mode="postpaid",
            assigned_rep=self.rep,
            is_active=True,
        )
        self.medicine = Medicine.objects.create(
            name="MR8 Med",
            pts=Decimal("100.00"),
            ptr=Decimal("90.00"),
            mrp=Decimal("120.00"),
            is_active=True,
        )
        DoctorMedicine.objects.create(doctor=self.doctor, medicine=self.medicine)

        # Build a settled campaign for most tests
        self.campaign = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=5,
            year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_SETTLED,
            total_sales_value=Decimal("1000.00"),
            total_commission=Decimal("100.00"),
            paid_amount=Decimal("100.00"),
        )

    # ── A. Legacy purge ────────────────────────────────────────────────────

    def test_postpaid_entry_model_does_not_exist(self):
        """
        PostpaidEntry must not be importable from sales.models.
        The class was physically removed in MR-8.0.
        """
        import sales.models as sm
        self.assertFalse(
            hasattr(sm, "PostpaidEntry"),
            "PostpaidEntry should have been removed from sales.models in MR-8.0",
        )

    # ── B. Ledger protection ───────────────────────────────────────────────

    def test_settled_campaign_cannot_be_deleted(self):
        """Settled campaigns must raise ValidationError on delete."""
        with self.assertRaises(ValidationError):
            self.campaign.delete()

    def test_locked_campaign_cannot_be_deleted(self):
        """Locked campaigns must raise ValidationError on delete."""
        locked = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=4,
            year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_LOCKED,
            total_sales_value=Decimal("500.00"),
            total_commission=Decimal("50.00"),
            paid_amount=Decimal("50.00"),
        )
        with self.assertRaises(ValidationError):
            locked.delete()

    def test_partial_campaign_cannot_be_deleted(self):
        """Partial campaigns must also raise ValidationError on delete."""
        partial = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=3,
            year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_PARTIAL,
            total_sales_value=Decimal("500.00"),
            total_commission=Decimal("50.00"),
            paid_amount=Decimal("25.00"),
        )
        with self.assertRaises(ValidationError):
            partial.delete()

    def test_open_campaign_can_be_deleted(self):
        """Open campaigns have no financial commitment — deletion is allowed."""
        open_campaign = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=2,
            year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_OPEN,
        )
        pk = open_campaign.pk
        open_campaign.delete()
        self.assertFalse(
            PostpaidCampaign.objects.filter(pk=pk).exists()
        )

    def test_sale_entry_on_settled_campaign_is_blocked(self):
        """
        PostpaidSaleEntry cannot be added to a Settled campaign.
        """
        with self.assertRaises(ValidationError):
            PostpaidSaleEntry.objects.create(
                campaign=self.campaign,
                medicine=self.medicine,
                quantity=5,
                entry_date=date.today(),
                rep=self.rep,
            )

    def test_campaign_payment_on_settled_campaign_is_blocked(self):
        """
        CampaignPayments can only be made on Partial campaigns;
        a Settled campaign must reject new payments.
        """
        with self.assertRaises(ValidationError):
            CampaignPayment.objects.create(
                campaign=self.campaign,
                amount=Decimal("10.00"),
                payment_date=date.today(),
            )

    def test_locked_campaign_is_immutable(self):
        """
        A Locked campaign must raise ValidationError on any attempted edit.
        """
        locked = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=6,
            year=2025,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_LOCKED,
        )
        locked.settlement_notes = "Attempting to mutate"
        with self.assertRaises(ValidationError):
            locked.save()

    # ── C. Correction model ────────────────────────────────────────────────

    def test_correction_creation_on_settled_campaign(self):
        """
        A correction should be creatable on a Settled campaign with all
        required fields, and snapshot values should be captured correctly.
        """
        correction = PostpaidCampaignCorrection(
            campaign=self.campaign,
            corrected_by=self.admin,
            correction_reason=PostpaidCampaignCorrection.REASON_WRITE_OFF,
            amount_adjustment=Decimal("-50.00"),
            notes="Writing off the balance as approved by management.",
            reference="MGMT-2026-001",
        )
        correction.save()

        self.assertIsNotNone(correction.pk)
        self.assertEqual(correction.snapshot_total_commission, Decimal("100.00"))
        self.assertEqual(correction.snapshot_paid_amount, Decimal("100.00"))
        self.assertEqual(correction.snapshot_outstanding_balance, Decimal("0.00"))

    def test_correction_on_locked_campaign(self):
        """Corrections should also work on Locked campaigns."""
        locked = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=1,
            year=2025,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_LOCKED,
            total_sales_value=Decimal("2000.00"),
            total_commission=Decimal("200.00"),
            paid_amount=Decimal("180.00"),
        )
        correction = PostpaidCampaignCorrection(
            campaign=locked,
            corrected_by=self.admin,
            correction_reason=PostpaidCampaignCorrection.REASON_PAYMENT_MISSED,
            amount_adjustment=Decimal("20.00"),
            notes="Payment of 20 was missed in ledger.",
        )
        correction.save()
        self.assertIsNotNone(correction.pk)
        # Snapshot should reflect locked campaign state
        self.assertEqual(correction.snapshot_outstanding_balance, Decimal("20.00"))

    def test_correction_on_open_campaign_is_blocked(self):
        """
        Corrections are only permitted on Settled or Locked campaigns.
        An Open campaign must be rejected.
        """
        open_campaign = PostpaidCampaign.objects.create(
            doctor=self.doctor,
            month=7,
            year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_OPEN,
        )
        correction = PostpaidCampaignCorrection(
            campaign=open_campaign,
            corrected_by=self.admin,
            correction_reason=PostpaidCampaignCorrection.REASON_OTHER,
            amount_adjustment=Decimal("-10.00"),
            notes="Invalid correction attempt.",
        )
        with self.assertRaises(ValidationError):
            correction.save()

    def test_correction_is_append_only(self):
        """
        Calling save() on an existing correction must raise ValueError.
        Corrections must never be mutated after creation.
        """
        correction = PostpaidCampaignCorrection.objects.create(
            campaign=self.campaign,
            corrected_by=self.admin,
            correction_reason=PostpaidCampaignCorrection.REASON_DATA_CORRECTION,
            amount_adjustment=Decimal("-5.00"),
            notes="Test append-only enforcement.",
        )
        correction.notes = "Trying to change notes after creation"
        with self.assertRaises(ValueError):
            correction.save()

    def test_correction_cannot_be_deleted(self):
        """
        delete() on a correction must raise ValidationError.
        Corrections form a permanent audit trail.
        """
        correction = PostpaidCampaignCorrection.objects.create(
            campaign=self.campaign,
            corrected_by=self.admin,
            correction_reason=PostpaidCampaignCorrection.REASON_MANAGEMENT_APPROVAL,
            amount_adjustment=Decimal("-20.00"),
            notes="Test delete protection.",
        )
        with self.assertRaises(ValidationError):
            correction.delete()
        # Verify record still exists in DB
        self.assertTrue(
            PostpaidCampaignCorrection.objects.filter(pk=correction.pk).exists()
        )

    def test_correction_requires_notes(self):
        """A correction with empty notes must fail validation."""
        correction = PostpaidCampaignCorrection(
            campaign=self.campaign,
            corrected_by=self.admin,
            correction_reason=PostpaidCampaignCorrection.REASON_OTHER,
            amount_adjustment=Decimal("-5.00"),
            notes="   ",  # whitespace only
        )
        with self.assertRaises(ValidationError) as ctx:
            correction.save()
        self.assertIn("notes", ctx.exception.message_dict)

    def test_correction_zero_adjustment_is_blocked(self):
        """A zero adjustment amount must fail validation."""
        correction = PostpaidCampaignCorrection(
            campaign=self.campaign,
            corrected_by=self.admin,
            correction_reason=PostpaidCampaignCorrection.REASON_OTHER,
            amount_adjustment=Decimal("0.00"),
            notes="Zero adjustment test.",
        )
        with self.assertRaises(ValidationError) as ctx:
            correction.save()
        self.assertIn("amount_adjustment", ctx.exception.message_dict)

    def test_correction_snapshot_independence(self):
        """
        The snapshot on a correction must not change even if the campaign
        totals are updated later.
        """
        correction = PostpaidCampaignCorrection.objects.create(
            campaign=self.campaign,
            corrected_by=self.admin,
            correction_reason=PostpaidCampaignCorrection.REASON_DATA_CORRECTION,
            amount_adjustment=Decimal("-10.00"),
            notes="Snapshot independence test.",
        )
        original_commission_snapshot = correction.snapshot_total_commission

        # Directly mutate campaign (bypassing guards for test purposes)
        PostpaidCampaign.objects.filter(pk=self.campaign.pk).update(
            total_commission=Decimal("999.00")
        )

        # Correction snapshot must be unchanged
        correction.refresh_from_db()
        self.assertEqual(
            correction.snapshot_total_commission,
            original_commission_snapshot,
        )

    def test_campaign_with_corrections_is_protected_from_deletion(self):
        """
        A Settled campaign that has corrections cannot be deleted
        (PROTECT FK on PostpaidCampaignCorrection.campaign).
        Even if we bypass the model-level delete guard, the DB constraint
        would prevent it — but here we test that the model guard fires first.
        """
        PostpaidCampaignCorrection.objects.create(
            campaign=self.campaign,
            corrected_by=self.admin,
            correction_reason=PostpaidCampaignCorrection.REASON_WRITE_OFF,
            amount_adjustment=Decimal("-50.00"),
            notes="Test FK protection.",
        )
        # Model-level guard fires before FK cascade
        with self.assertRaises(ValidationError):
            self.campaign.delete()


# ─────────────────────────────────────────────────────────────────────────────
# MR-9.0: Postpaid Report Separation
# ─────────────────────────────────────────────────────────────────────────────


class PostpaidReportTests(TestCase):
    """
    Tests for MR-9.0 Postpaid Report Separation:
    A. Postpaid Sales Report aggregation accuracy.
    B. Filter correctness (doctor, rep, month, year, status).
    C. Status handling in sales report.
    D. Settlement Ledger Report calculations.
    E. Export generation (Excel bytes produced without error).
    """

    def setUp(self):
        self.admin = User.objects.create_user(
            username="mr9_admin", password="pwd", role="admin"
        )
        self.rep1 = User.objects.create_user(
            username="mr9_rep1", password="pwd", role="rep",
            first_name="Alice", last_name="Smith",
        )
        self.rep2 = User.objects.create_user(
            username="mr9_rep2", password="pwd", role="rep",
            first_name="Bob", last_name="Jones",
        )
        self.doctor_a = Doctor.objects.create(
            name="Dr. Alpha", mode="postpaid",
            assigned_rep=self.rep1, is_active=True,
        )
        self.doctor_b = Doctor.objects.create(
            name="Dr. Beta", mode="postpaid",
            assigned_rep=self.rep2, is_active=True,
        )
        self.med_x = Medicine.objects.create(
            name="Med X", pts=Decimal("100.00"),
            ptr=Decimal("90.00"), mrp=Decimal("120.00"), is_active=True,
        )
        self.med_y = Medicine.objects.create(
            name="Med Y", pts=Decimal("200.00"),
            ptr=Decimal("180.00"), mrp=Decimal("240.00"), is_active=True,
        )
        DoctorMedicine.objects.create(doctor=self.doctor_a, medicine=self.med_x)
        DoctorMedicine.objects.create(doctor=self.doctor_a, medicine=self.med_y)
        DoctorMedicine.objects.create(doctor=self.doctor_b, medicine=self.med_x)

        # Campaign A1 — Dr. Alpha, Jan 2026, 10%, Open
        self.camp_a1 = PostpaidCampaign.objects.create(
            doctor=self.doctor_a, month=1, year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_OPEN,
        )
        # 2 entries: 3 × MedX + 2 × MedY
        self.sale_a1_x = PostpaidSaleEntry.objects.create(
            campaign=self.camp_a1, medicine=self.med_x, quantity=3,
            entry_date=date(2026, 1, 5), rep=self.rep1,
        )
        self.sale_a1_y = PostpaidSaleEntry.objects.create(
            campaign=self.camp_a1, medicine=self.med_y, quantity=2,
            entry_date=date(2026, 1, 10), rep=self.rep1,
        )

        # Campaign B1 — Dr. Beta, Feb 2026.
        # Must be Open when sale entry is created; then forced to Partial via
        # queryset.update() (bypasses model-level clean() guard — intentional
        # for test setup only, not a pattern used in production code).
        self.camp_b1 = PostpaidCampaign.objects.create(
            doctor=self.doctor_b, month=2, year=2026,
            commission_percentage=Decimal("15.00"),
            status=PostpaidCampaign.STATUS_OPEN,
        )
        self.sale_b1_x = PostpaidSaleEntry.objects.create(
            campaign=self.camp_b1, medicine=self.med_x, quantity=5,
            entry_date=date(2026, 2, 15), rep=self.rep2,
        )
        # Advance camp_b1 to Partial — queryset.update() skips clean()
        PostpaidCampaign.objects.filter(pk=self.camp_b1.pk).update(
            status=PostpaidCampaign.STATUS_PARTIAL,
            total_sales_value=Decimal("500.00"),
            total_commission=Decimal("75.00"),
            paid_amount=Decimal("30.00"),
        )
        self.camp_b1.refresh_from_db()

        # Settled campaign — ledger-only, no sale entries needed
        self.camp_settled = PostpaidCampaign.objects.create(
            doctor=self.doctor_a, month=3, year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_SETTLED,
            total_sales_value=Decimal("1000.00"),
            total_commission=Decimal("100.00"),
            paid_amount=Decimal("100.00"),
            settlement_reason=PostpaidCampaign.REASON_WRITE_OFF,
            settlement_notes="Full settlement.",
        )

    # ── A. Aggregation accuracy ───────────────────────────────────────────

    def test_postpaid_sales_summary_totals(self):
        """
        Summary totals must sum value_at_sale and commission_at_sale
        across all PostpaidSaleEntry rows.
        """
        qs = get_postpaid_sales_queryset()
        summary = get_postpaid_sales_summary(qs)

        # Expected values:
        # sale_a1_x: 3 × 100 = 300, comm = 300 × 10% = 30
        # sale_a1_y: 2 × 200 = 400, comm = 400 × 10% = 40
        # sale_b1_x: 5 × 100 = 500, comm = 500 × 15% = 75
        self.assertEqual(summary["total_entries"], 3)
        self.assertEqual(summary["total_quantity"], 10)
        self.assertEqual(summary["total_value"], Decimal("1200.00"))
        self.assertEqual(summary["total_commission"], Decimal("145.00"))

    def test_postpaid_sales_report_row_structure(self):
        """Each row must have all required keys with correct values."""
        qs = get_postpaid_sales_queryset(doctor_id=self.doctor_a.id)
        rows = get_postpaid_sales_report(qs)

        self.assertEqual(len(rows), 2)
        row = rows[0]  # ordered by -year/-month then -entry_date, so sale_a1_y first
        required_keys = [
            "sl_no", "period", "month", "year", "doctor_name", "rep_name",
            "medicine_name", "quantity", "pts_at_sale", "value_at_sale",
            "commission_pct", "commission_earned", "entry_date", "campaign_status",
        ]
        for key in required_keys:
            self.assertIn(key, row, f"Missing key: {key}")

    # ── B. Filter correctness ───────────────────────────────────────────

    def test_filter_by_doctor(self):
        qs = get_postpaid_sales_queryset(doctor_id=self.doctor_a.id)
        self.assertEqual(qs.count(), 2)  # sale_a1_x + sale_a1_y

    def test_filter_by_rep(self):
        qs = get_postpaid_sales_queryset(rep_id=self.rep2.id)
        self.assertEqual(qs.count(), 1)  # only sale_b1_x

    def test_filter_by_month(self):
        qs = get_postpaid_sales_queryset(month=1)
        self.assertEqual(qs.count(), 2)  # January entries only

    def test_filter_by_year(self):
        qs = get_postpaid_sales_queryset(year=2026)
        self.assertEqual(qs.count(), 3)  # all 3 in 2026

    def test_filter_by_status(self):
        qs = get_postpaid_sales_queryset(status=PostpaidCampaign.STATUS_PARTIAL)
        self.assertEqual(qs.count(), 1)  # only sale_b1_x (camp_b1 is partial)

    def test_filter_combined_doctor_and_month(self):
        qs = get_postpaid_sales_queryset(doctor_id=self.doctor_a.id, month=1)
        self.assertEqual(qs.count(), 2)

    def test_filter_no_results(self):
        qs = get_postpaid_sales_queryset(month=12, year=2026)
        self.assertEqual(qs.count(), 0)
        summary = get_postpaid_sales_summary(qs)
        self.assertEqual(summary["total_entries"], 0)
        self.assertEqual(summary["total_value"], Decimal("0"))

    # ── C. Status handling ───────────────────────────────────────────

    def test_status_display_in_rows(self):
        """
        campaign_status in each row should reflect the readable display
        value from get_status_display().
        """
        qs = get_postpaid_sales_queryset(doctor_id=self.doctor_a.id)
        rows = get_postpaid_sales_report(qs)
        statuses = {r["campaign_status"] for r in rows}
        # camp_a1 is Open
        self.assertIn("Open", statuses)

    def test_settled_status_in_rows(self):
        """
        Sales report rows from Open campaigns show 'Open'.
        Settled status is verified via the settlement ledger (which does not
        require inserting entries into a protected campaign).
        """
        qs   = get_postpaid_sales_queryset(doctor_id=self.doctor_a.id)
        rows = get_postpaid_sales_report(qs)
        statuses = {r["campaign_status"] for r in rows}
        # camp_a1 is Open
        self.assertIn("Open", statuses)

        # Verify 'settled' appears in the ledger report status field
        ledger_qs   = get_settlement_ledger_queryset(status=PostpaidCampaign.STATUS_SETTLED)
        ledger_rows = get_settlement_ledger_report(ledger_qs)
        self.assertEqual(len(ledger_rows), 1)
        self.assertEqual(ledger_rows[0]["status"], PostpaidCampaign.STATUS_SETTLED)
        self.assertTrue(ledger_rows[0]["is_settled"])

    # ── D. Settlement Ledger ───────────────────────────────────────────

    def test_settlement_ledger_row_count(self):
        """Ledger must return one row per PostpaidCampaign."""
        qs   = get_settlement_ledger_queryset()
        rows = get_settlement_ledger_report(qs)
        # 3 campaigns: camp_a1, camp_b1, camp_settled
        self.assertEqual(len(rows), 3)

    def test_settlement_ledger_outstanding_balance(self):
        """
        Outstanding balance in ledger rows must equal
        total_commission - paid_amount.
        """
        qs   = get_settlement_ledger_queryset(doctor_id=self.doctor_b.id)
        rows = get_settlement_ledger_report(qs)
        self.assertEqual(len(rows), 1)
        row = rows[0]
        # camp_b1: 75 commission, 30 paid → 45 outstanding
        self.assertEqual(row["total_commission"],    Decimal("75.00"))
        self.assertEqual(row["paid_amount"],          Decimal("30.00"))
        self.assertEqual(row["outstanding_balance"],  Decimal("45.00"))

    def test_settlement_summary_aggregation(self):
        """Summary must aggregate all ledger campaigns correctly."""
        qs      = get_settlement_ledger_queryset()
        summary = get_settlement_summary(qs)
        self.assertEqual(summary["total_campaigns"], 3)
        self.assertEqual(summary["settled_count"],   1)
        self.assertEqual(summary["partial_count"],   1)
        self.assertEqual(summary["open_count"],      1)

    def test_settlement_filter_by_status_settled(self):
        """Filtering ledger by 'settled' must return only settled campaigns."""
        qs   = get_settlement_ledger_queryset(status=PostpaidCampaign.STATUS_SETTLED)
        rows = get_settlement_ledger_report(qs)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["doctor_name"], "Dr. Alpha")
        self.assertTrue(rows[0]["is_settled"])

    def test_settlement_filter_by_doctor(self):
        """Doctor filter must narrow ledger to that doctor's campaigns."""
        qs   = get_settlement_ledger_queryset(doctor_id=self.doctor_a.id)
        rows = get_settlement_ledger_report(qs)
        # Dr. Alpha has camp_a1 (Open) + camp_settled (Settled) = 2
        self.assertEqual(len(rows), 2)

    def test_settlement_filter_by_month_and_year(self):
        qs   = get_settlement_ledger_queryset(month=2, year=2026)
        rows = get_settlement_ledger_report(qs)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["period"], "02/2026")

    def test_settlement_locked_flag(self):
        """is_locked must be True only for Locked campaigns."""
        locked_camp = PostpaidCampaign.objects.create(
            doctor=self.doctor_b, month=4, year=2026,
            commission_percentage=Decimal("10.00"),
            status=PostpaidCampaign.STATUS_LOCKED,
            total_commission=Decimal("50.00"),
            paid_amount=Decimal("50.00"),
        )
        qs   = get_settlement_ledger_queryset(doctor_id=self.doctor_b.id)
        rows = get_settlement_ledger_report(qs)
        locked_rows = [r for r in rows if r["is_locked"]]
        self.assertEqual(len(locked_rows), 1)
        self.assertEqual(locked_rows[0]["status"], PostpaidCampaign.STATUS_LOCKED)

    # ── E. Export generation ──────────────────────────────────────────

    def test_postpaid_sales_excel_export_produces_bytes(self):
        """
        export_postpaid_sales_to_excel must return a non-empty BytesIO buffer
        that is a valid OpenPyXL-readable workbook.
        """
        from io import BytesIO
        from openpyxl import load_workbook

        qs      = get_postpaid_sales_queryset()
        summary = get_postpaid_sales_summary(qs)
        rows    = get_postpaid_sales_report(qs)
        summary["_rows"]            = rows

        buf = export_postpaid_sales_to_excel(qs, summary)
        self.assertIsInstance(buf, BytesIO)
        self.assertGreater(len(buf.getvalue()), 0)

        buf.seek(0)
        wb = load_workbook(buf)
        # 1 sheet expected
        self.assertEqual(len(wb.sheetnames), 1)
        self.assertIn("Sales Detail", wb.sheetnames)

    def test_postpaid_sales_excel_data_rows(self):
        """
        The Sales Detail sheet must contain one data row per PostpaidSaleEntry
        that matches the filter. We scope to this test's doctors to isolate
        from entries created by other test classes in the same test run.
        """
        from io import BytesIO
        from openpyxl import load_workbook

        # Scope to doctor_a and doctor_b only — 3 entries total
        doctor_ids = [self.doctor_a.id, self.doctor_b.id]
        qs = PostpaidSaleEntry.objects.filter(
            campaign__doctor_id__in=doctor_ids
        ).select_related("campaign", "campaign__doctor", "medicine", "rep")

        from sales.services.postpaid_service import (
            get_postpaid_sales_report as _report,
            get_postpaid_sales_summary as _summary,
        )
        summary = _summary(qs)
        rows    = _report(qs)
        for k in ["_rows"]:
            summary[k] = []
        summary["_rows"] = rows

        buf = export_postpaid_sales_to_excel(rows, summary)
        buf.seek(0)
        ws = load_workbook(buf)["Sales Detail"]
        # Data rows have an integer Sl No in column 1.
        # The summary block below uses text labels, so we filter by int.
        data_rows = [
            r for r in ws.iter_rows(min_row=4, values_only=True)
            if r[0] is not None and isinstance(r[0], int)
        ]
        # Sheet must contain exactly as many rows as we passed in
        self.assertEqual(len(data_rows), len(rows))
        self.assertGreaterEqual(len(data_rows), 3)

    def test_settlement_ledger_excel_export_produces_bytes(self):
        """
        export_settlement_ledger_to_excel must return a non-empty BytesIO
        readable by OpenPyXL with the Settlement Ledger sheet.
        We scope to this test's doctors to isolate from other test classes.
        """
        from io import BytesIO
        from openpyxl import load_workbook

        doctor_ids = [self.doctor_a.id, self.doctor_b.id]
        qs      = get_settlement_ledger_queryset()
        qs      = qs.filter(doctor_id__in=doctor_ids)
        rows    = get_settlement_ledger_report(qs)
        summary = get_settlement_summary(qs)

        buf = export_settlement_ledger_to_excel(rows, summary)
        self.assertIsInstance(buf, BytesIO)
        self.assertGreater(len(buf.getvalue()), 0)

        buf.seek(0)
        wb = load_workbook(buf)
        self.assertIn("Settlement Ledger", wb.sheetnames)
        ws = wb["Settlement Ledger"]
        # Data rows have integer Sl No in column 1.
        data_rows = [
            r for r in ws.iter_rows(min_row=4, values_only=True)
            if r[0] is not None and isinstance(r[0], int)
        ]
        self.assertEqual(len(data_rows), len(rows))
        self.assertGreaterEqual(len(data_rows), 3)

    def test_empty_queryset_export_does_not_raise(self):
        """
        Export functions must not raise when given an empty queryset.
        """
        from io import BytesIO
        from openpyxl import load_workbook

        qs      = get_postpaid_sales_queryset(month=12, year=2025)  # no data
        summary = get_postpaid_sales_summary(qs)
        rows    = get_postpaid_sales_report(qs)
        for k in ["_rows"]:
            summary[k] = []

        buf = export_postpaid_sales_to_excel(qs, summary)
        buf.seek(0)
        wb = load_workbook(buf)
        self.assertIn("Sales Detail", wb.sheetnames)

        qs2      = get_settlement_ledger_queryset(month=12, year=2025)
        rows2    = get_settlement_ledger_report(qs2)
        summary2 = get_settlement_summary(qs2)
        buf2 = export_settlement_ledger_to_excel(rows2, summary2)
        buf2.seek(0)
        wb2 = load_workbook(buf2)
        self.assertIn("Settlement Ledger", wb2.sheetnames)

    def test_campaign_monitor_view_shows_all(self):
        """
        The Campaign Monitor (postpaid_report_view) must show all campaigns
        regardless of status.
        """
        self.client.force_login(self.admin)
        from django.urls import reverse
        response = self.client.get(reverse("sales:postpaid_report"))
        self.assertEqual(response.status_code, 200)
        campaigns = response.context["campaigns"]
        statuses = {c.status for c in campaigns}
        # camp_a1 (Open), camp_b1 (Partial) are active
        self.assertIn(PostpaidCampaign.STATUS_OPEN, statuses)
        self.assertIn(PostpaidCampaign.STATUS_PARTIAL, statuses)
        # camp_settled (Settled) should ALSO be included
        self.assertIn(PostpaidCampaign.STATUS_SETTLED, statuses)

    def test_settlement_ledger_view_filters_closed(self):
        """
        The Settlement Ledger view must only show closed campaigns (Settled, Locked)
        and filter out active ones.
        """
        self.client.force_login(self.admin)
        from django.urls import reverse
        response = self.client.get(reverse("sales:settlement_ledger"))
        self.assertEqual(response.status_code, 200)
        rows = response.context["rows"]
        statuses = {r["status"] for r in rows}
        # Settled should be present
        self.assertIn(PostpaidCampaign.STATUS_SETTLED, statuses)
        # Open and Partial (active) should be excluded
        self.assertNotIn(PostpaidCampaign.STATUS_OPEN, statuses)
        self.assertNotIn(PostpaidCampaign.STATUS_PARTIAL, statuses)

    def test_export_views_return_ok(self):
        """
        The export views must return Excel file responses without crashes.
        """
        self.client.force_login(self.admin)
        from django.urls import reverse
        # Export sales
        response = self.client.get(reverse("sales:export_postpaid_sales"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Export ledger
        response = self.client.get(reverse("sales:export_settlement_ledger"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


class CriticalProductionFixesTests(TestCase):
    def setUp(self):
        self.rep1 = User.objects.create_user(username="rep1_fixes", password="pwd", role="rep")
        self.rep2 = User.objects.create_user(username="rep2_fixes", password="pwd", role="rep")
        self.admin = User.objects.create_user(username="admin_fixes", password="pwd", role="admin")
        self.admin.is_superuser = True
        self.admin.is_staff = True
        self.admin.save()

        self.doctor1 = Doctor.objects.create(
            name="Dr. Assigned to Rep1",
            mode="postpaid",
            assigned_rep=self.rep1,
            is_active=True
        )
        self.medicine = Medicine.objects.create(
            name="Medicine Fixes",
            pts=Decimal("120.00"),
            ptr=Decimal("100.00"),
            mrp=Decimal("150.00"),
            is_active=True
        )
        DoctorMedicine.objects.create(doctor=self.doctor1, medicine=self.medicine)

    def test_cross_rep_api_access(self):
        """
        Reps should be forbidden (403) from accessing another rep's doctor data via AJAX endpoints.
        """
        # Log in as rep2 (who is NOT assigned to doctor1)
        self.client.force_login(self.rep2)
        from django.urls import reverse

        # 1. Test api_medicines_for_doctor
        url_meds = reverse("sales:api_medicines", kwargs={"doctor_id": self.doctor1.id})
        response = self.client.get(url_meds)
        self.assertEqual(response.status_code, 403)

        # 2. Test api_campaign_for_doctor
        url_camp = reverse("sales:api_campaign", kwargs={"doctor_id": self.doctor1.id, "month": 6, "year": 2026})
        response = self.client.get(url_camp)
        self.assertEqual(response.status_code, 403)

        # Log in as rep1 (assigned rep) - should succeed (200)
        self.client.force_login(self.rep1)
        response = self.client.get(url_meds)
        self.assertEqual(response.status_code, 200)

        response = self.client.get(url_camp)
        self.assertEqual(response.status_code, 200)

    def test_admin_api_access(self):
        """
        Admins should be allowed (200) to access any doctor's data via AJAX endpoints.
        """
        self.client.force_login(self.admin)
        from django.urls import reverse

        url_meds = reverse("sales:api_medicines", kwargs={"doctor_id": self.doctor1.id})
        response = self.client.get(url_meds)
        self.assertEqual(response.status_code, 200)

        url_camp = reverse("sales:api_campaign", kwargs={"doctor_id": self.doctor1.id, "month": 6, "year": 2026})
        response = self.client.get(url_camp)
        self.assertEqual(response.status_code, 200)

    def test_campaign_delete_action_removed(self):
        """
        Verify that bulk delete_selected action is removed from PostpaidCampaignAdmin.
        """
        from django.contrib.admin.sites import site
        from django.test.client import RequestFactory

        admin_inst = site._registry[PostpaidCampaign]
        request = RequestFactory().get('/admin/sales/postpaidcampaign/')
        request.user = self.admin
        actions = admin_inst.get_actions(request)
        self.assertNotIn('delete_selected', actions)

    def test_locked_campaign_delete_denied(self):
        """
        Verify that PostpaidCampaignAdmin delete permission is allowed only for Open status,
        and denied for Awaiting Commission, Partial, Settled, and Locked campaigns.
        """
        from django.contrib.admin.sites import site
        from django.test.client import RequestFactory

        admin_inst = site._registry[PostpaidCampaign]
        request = RequestFactory().get('/admin/sales/postpaidcampaign/')
        request.user = self.admin

        # Open status campaign -> deletable (True)
        camp_open = PostpaidCampaign.objects.create(
            doctor=self.doctor1, month=1, year=2026, status=PostpaidCampaign.STATUS_OPEN, commission_percentage=Decimal("10.00")
        )
        self.assertTrue(admin_inst.has_delete_permission(request, camp_open))

        # Awaiting Commission status campaign -> NOT deletable (False)
        camp_awaiting = PostpaidCampaign.objects.create(
            doctor=self.doctor1, month=2, year=2026, status=PostpaidCampaign.STATUS_AWAITING_COMMISSION
        )
        self.assertFalse(admin_inst.has_delete_permission(request, camp_awaiting))

        # Partial status campaign -> NOT deletable (False)
        camp_partial = PostpaidCampaign.objects.create(
            doctor=self.doctor1, month=3, year=2026, status=PostpaidCampaign.STATUS_PARTIAL, commission_percentage=Decimal("10.00")
        )
        self.assertFalse(admin_inst.has_delete_permission(request, camp_partial))

        # Settled status campaign -> NOT deletable (False)
        camp_settled = PostpaidCampaign.objects.create(
            doctor=self.doctor1, month=4, year=2026, status=PostpaidCampaign.STATUS_SETTLED, commission_percentage=Decimal("10.00")
        )
        self.assertFalse(admin_inst.has_delete_permission(request, camp_settled))

        # Locked status campaign -> NOT deletable (False)
        camp_locked = PostpaidCampaign.objects.create(
            doctor=self.doctor1, month=5, year=2026, status=PostpaidCampaign.STATUS_LOCKED, commission_percentage=Decimal("10.00")
        )
        self.assertFalse(admin_inst.has_delete_permission(request, camp_locked))

    def test_atomic_transaction_protection(self):
        """
        Verify that postpaid_sales_entry_view uses select_for_update to lock campaign.
        """
        from unittest.mock import patch
        from django.db.models.query import QuerySet
        from django.urls import reverse

        self.client.force_login(self.rep1)
        url = reverse("sales:postpaid_entry")

        original_select_for_update = QuerySet.select_for_update

        with patch.object(QuerySet, 'select_for_update', autospec=True) as mock_select:
            mock_select.side_effect = lambda self, *args, **kwargs: original_select_for_update(self, *args, **kwargs)
            
            response = self.client.post(url, {
                "doctor": self.doctor1.id,
                "month": 6,
                "year": 2026,
                "medicine": self.medicine.id,
                "quantity": 10,
                "notes": "Testing lock"
            })
            
            # Post request should succeed or redirect, and mock_select must have been called
            self.assertEqual(response.status_code, 302)
            self.assertTrue(mock_select.called)


