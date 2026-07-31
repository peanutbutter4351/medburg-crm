from decimal import Decimal
from datetime import date, timedelta
from django.test import TestCase
from django.contrib.auth import get_user_model
from django.urls import reverse

from doctors.models import Doctor
from medicines.models import Medicine
from sales.models import PostpaidCampaign, PostpaidSaleEntry
from reports.services.postpaid_report_service import (
    get_postpaid_doctor_report_queryset,
    get_postpaid_doctor_report_summary,
)

User = get_user_model()


class PostpaidDoctorReportTests(TestCase):
    def setUp(self):
        # Create users
        self.admin = User.objects.create_user(username="admin", role="admin")
        self.rep = User.objects.create_user(username="rep", role="rep")

        # Create medicines
        self.med1 = Medicine.objects.create(
            name="Medicine A",
            brand="Brand A",
            pts=Decimal("100.00"),
            ptr=Decimal("120.00"),
            mrp=Decimal("150.00"),
            is_active=True,
        )
        self.med2 = Medicine.objects.create(
            name="Medicine B",
            brand="Brand B",
            pts=Decimal("200.00"),
            ptr=Decimal("240.00"),
            mrp=Decimal("300.00"),
            is_active=True,
        )

        # Create doctors
        self.doctor1 = Doctor.objects.create(
            name="Doctor John",
            mode="postpaid",
            location="New York",
            is_active=True,
        )
        self.doctor2 = Doctor.objects.create(
            name="Doctor Alice",
            mode="postpaid",
            location="Boston",
            is_active=True,
        )

        # Create campaigns
        # Doctor 1: Jan Campaign
        self.camp1 = PostpaidCampaign.objects.create(
            doctor=self.doctor1,
            month=1,
            year=2026,
            commission_percentage=Decimal("10.0"),
            status=PostpaidCampaign.STATUS_OPEN,
        )
        # Doctor 1: Feb Campaign
        self.camp2 = PostpaidCampaign.objects.create(
            doctor=self.doctor1,
            month=2,
            year=2026,
            commission_percentage=Decimal("15.0"),
            status=PostpaidCampaign.STATUS_OPEN,
        )
        # Doctor 2: Jan Campaign (created as OPEN first to allow sales entries)
        self.camp3 = PostpaidCampaign.objects.create(
            doctor=self.doctor2,
            month=1,
            year=2026,
            commission_percentage=Decimal("5.0"),
            status=PostpaidCampaign.STATUS_OPEN,
        )

        # Create Sales Entries
        # Doctor 1 Jan (Medicine A): Quantity 10 -> Value 1000, Comm 100
        self.sale1 = PostpaidSaleEntry.objects.create(
            campaign=self.camp1,
            medicine=self.med1,
            quantity=10,
            entry_date=date(2026, 1, 15),
            rep=self.rep,
        )
        # Doctor 1 Feb (Medicine B): Quantity 5 -> Value 1000, Comm 150
        self.sale2 = PostpaidSaleEntry.objects.create(
            campaign=self.camp2,
            medicine=self.med2,
            quantity=5,
            entry_date=date(2026, 2, 10),
            rep=self.rep,
        )
        # Doctor 2 Jan (Medicine A): Quantity 20 -> Value 2000, Comm 100
        self.sale3 = PostpaidSaleEntry.objects.create(
            campaign=self.camp3,
            medicine=self.med1,
            quantity=20,
            entry_date=date(2026, 1, 20),
            rep=self.rep,
        )

        # Now save camp3 as settled
        self.camp3.status = PostpaidCampaign.STATUS_SETTLED
        self.camp3.settlement_reason = PostpaidCampaign.REASON_WRITE_OFF
        self.camp3.settlement_notes = "settled under test waiver"
        self.camp3.save(update_fields=["status", "settlement_reason", "settlement_notes"])

    def test_access_control(self):
        """Verify report page and export are restricted to admin only."""
        report_url = reverse("reports:postpaid_doctor_report")
        export_url = reverse("reports:export_postpaid_doctor_report")

        # Rep (Non-admin) -> 403 Forbidden
        self.client.force_login(self.rep)
        response = self.client.get(report_url)
        self.assertEqual(response.status_code, 403)
        response_export = self.client.get(export_url)
        self.assertEqual(response_export.status_code, 403)

        # Admin -> 200 OK
        self.client.force_login(self.admin)
        response = self.client.get(report_url)
        self.assertEqual(response.status_code, 200)
        response_export = self.client.get(export_url)
        self.assertEqual(response_export.status_code, 200)

    def test_aggregation_without_filters(self):
        """Verify consolidated totals across multiple campaigns per doctor."""
        qs = get_postpaid_doctor_report_queryset()
        self.assertEqual(qs.count(), 2)

        # Alice should be first in alphabetical order (A-Z)
        alice = qs[0]
        self.assertEqual(alice.name, "Doctor Alice")
        self.assertEqual(alice.total_sales_value, Decimal("2000.00"))
        self.assertEqual(alice.total_commission, Decimal("100.00"))

        # John is second
        john = qs[1]
        self.assertEqual(john.name, "Doctor John")
        self.assertEqual(john.total_sales_value, Decimal("2000.00"))  # camp1 (1000) + camp2 (1000)
        self.assertEqual(john.total_commission, Decimal("250.00"))    # camp1 (100) + camp2 (150)

        # Summary
        summary = get_postpaid_doctor_report_summary(qs)
        self.assertEqual(summary["total_doctors"], 2)
        self.assertEqual(summary["total_sales"], Decimal("4000.00"))
        self.assertEqual(summary["total_commission"], Decimal("350.00"))
        self.assertEqual(summary["average_commission_pct"], Decimal("8.75"))  # (350 / 4000) * 100

    def test_filters(self):
        """Verify that status, medicine, location, doctor, and date range filters work."""
        # 1. Filter by Doctor
        qs = get_postpaid_doctor_report_queryset(doctor_id=self.doctor1.id)
        self.assertEqual(qs.count(), 1)
        self.assertEqual(qs[0].name, "Doctor John")

        # 2. Filter by Location
        qs = get_postpaid_doctor_report_queryset(location="Boston")
        self.assertEqual(qs.count(), 1)
        self.assertEqual(qs[0].name, "Doctor Alice")

        # 3. Filter by Status (STATUS_SETTLED)
        qs = get_postpaid_doctor_report_queryset(status=PostpaidCampaign.STATUS_SETTLED)
        self.assertEqual(qs.count(), 1)
        self.assertEqual(qs[0].name, "Doctor Alice")

        # 4. Filter by Product/Medicine (Medicine B only on camp2)
        qs = get_postpaid_doctor_report_queryset(medicine_id=self.med2.id)
        self.assertEqual(qs.count(), 1)
        self.assertEqual(qs[0].name, "Doctor John")
        self.assertEqual(qs[0].total_sales_value, Decimal("1000.00"))
        self.assertEqual(qs[0].total_commission, Decimal("150.00"))

        # 5. Filter by Date Range (Jan 1st to Jan 31st)
        qs = get_postpaid_doctor_report_queryset(from_date=date(2026, 1, 1), to_date=date(2026, 1, 31))
        self.assertEqual(qs.count(), 2)  # both have sales in Jan
        john = qs.filter(pk=self.doctor1.pk).first()
        self.assertEqual(john.total_sales_value, Decimal("1000.00"))  # Feb sale excluded

    def test_empty_state(self):
        """Verify report page shows empty state messaging when filters yield no matches."""
        self.client.force_login(self.admin)
        response = self.client.get(reverse("reports:postpaid_doctor_report"), {"location": "InvalidLocation"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No postpaid doctor records match the selected filters.")

    def test_excel_export_header_and_data(self):
        """Verify the Excel download format and standard header presence."""
        self.client.force_login(self.admin)
        response = self.client.get(reverse("reports:export_postpaid_doctor_report"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Load workbook from response bytes
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check standard metadata header
        self.assertEqual(ws.cell(row=1, column=1).value, "Medburg Medical Products")
        self.assertEqual(ws.cell(row=2, column=1).value, "Postpaid Doctors Report")
        self.assertTrue(ws.cell(row=3, column=1).value.startswith("Generated:"))
        self.assertEqual(ws.cell(row=5, column=1).value, "Applied Filters:")
